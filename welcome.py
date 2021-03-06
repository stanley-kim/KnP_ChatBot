#-*- coding: utf-8 -*-
# Copyright 2015 IBM Corp. All Rights Reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import os
import json
import time
from flask import Flask, jsonify, request
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import smtplib
from datetime import datetime, timedelta
from threading import Timer

import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

import re

import traceback

app = Flask(__name__)

VersionString = u'1.21'


_State0KeyList = [ 
    u'1.고장 접수',   
    u'2.고장 접수 내역 확인',    
    u'3.사용방법 안내', 
    u'4.기타'
    ]

_State1KeyList = [
      u'1.지하4층 실습실',
      u'2.지하4층 기공실',
      u'3.지하3층 실습실(자리)',
      u'4.지하3층 실습실(핸드피스)',
      u'5.지하3층 컴퓨터실',
      u'이전 메뉴'
    ]   

_State_ItemListCheckTop_List = [
      u'이전 메뉴' ,
      u'1.일부 삭제',
      u'2.전체 삭제',
      u'이전 메뉴'     
    ]   
  
_State4KeyList = [
      u'1.학번(혹은 사번) 및 이름 입력/수정',
      u'2.학번(혹은 사번) 및 이름 삭제' ,
      u'3.입력 방식 변경' , 
      u'4.Forwarding 신청' , 
      u'5.시스템 관리자 메뉴' ,
      u'이전 메뉴'     
    ] 


_GradeKeyList = [
      u'1.본과 4학년',
      u'2.본과 3학년',
      u'3.본과 2학년',
      u'4.본과 1학년',
      u'5.예과 2학년',
      u'6.예과 1학년',
      u'이전 메뉴'                
]

_InputModeList = [
      u'1.버튼 방식',
      u'2.직접 입력 방식' ,
      u'3.혼합 입력 방식' ,
      u'이전 메뉴'          
]

_YesorNoKeyList = [  u'1.Yes', u'2.No' , u'이전 메뉴'  ]

_YesorNoKeyListv2 = [  u'1.Yes', u'2.Yes(+파트 추가 입력)', u'3.No' , u'이전 메뉴' , u'4.Yes(연습용)', u'5.Yes(+파트 추가 입력)(연습용)', u'6.No(연습용)' ]


_State13KeyList = [
      u'5.지하 3층 실습실 자리',
      u'6.지하 3층용 핸드피스',
      u'이전 메뉴'     
]

_State111KeyList = [
      u'라이트',
      u'모니터',
      u'가스토치',
      u'핸드피스 엔진',
      u'Air Suction(흡입구)',
      u'Air Spray(배출구)' ,
      u'직접 입력'  ,    
      u'이전 메뉴'     
    ]   

len_4work_part = len( _State111KeyList)

_State141KeyList = [
      u'모니터',
      u'본체',  
      u'네트워크',    
      u'직접 입력',    
      u'이전 메뉴'                
]
len_3com_part = len( _State141KeyList)


_State1311KeyList = [
      u'토마스',
      u'모니터',
      u'라이트',      
      u'3Way Air Water Syringe',   
      u'하이스피드 핸드피스 Connector'  ,
      u'로스피드 핸드피스 Connector'  ,               
      u'직접 입력'  ,    
      u'이전 메뉴'              
]
len_3work_part = len( _State1311KeyList)


_State1321KeyList = [
      u'하이스피드 핸드피스',
      u'로스피드 핸드피스',
      u'직접 입력'  ,    
      u'이전 메뉴'              
]
len_3handpiece_part = len( _State1321KeyList)


_LightSymptomKeyList = [
      u'1:안 켜짐',
      u'2:깜빡깜빡 거림',
      u'3:위치 고정 안 됨',
      u'4:증상 직접 입력' ,
      u'이전 메뉴'              
    ]   

_LightSymptomMultiChoiceList = [
      u'1:안 켜짐',
      u'2:깜빡깜빡 거림',
      u'3:위치 고정 안 됨'
    ]   

_LightSymptomJoinString = u'\n'.join(_LightSymptomMultiChoiceList)


_MonitorSymptomKeyList = [
      u'1:안 켜짐',
      u'2:깜빡깜빡 거림',
      u'3:흑백으로 나옴' ,
      u'4:증상 직접 입력'  ,
      u'이전 메뉴'         
    ]   


_MonitorSymptomMultiChoiceList = [
      u'1:안 켜짐',
      u'2:깜빡깜빡 거림',
      u'3:흑백으로 나옴' 
    ]   

_MonitorSymptomJoinString = u'\n'.join(_MonitorSymptomMultiChoiceList)

_110VoltSymptomKeyList = [
      u'1.전원 안 들어옴',
      u'2.구망에 안 들어감',
      u'3.구망에 들어가서 안 나옴',
      u'4.증상 직접 입력' ,
      u'이전 메뉴'         
]

_GastorchSymptomKeyList = [
      u'1:불 안 나옴',
      u'2:불 너무 약함',
      u'3:증상 직접 입력' ,
      u'이전 메뉴'    
]
_GastorchSymptomMultiChoiceList = [
      u'1:불 안 나옴',
      u'2:불 너무 약함'
]
_GastorchSymptomJoinString = u'\n'.join(_GastorchSymptomMultiChoiceList)



_220VoltSymptomKeyList = _110VoltSymptomKeyList

_HandpieceengineSymptomKeyList = [
      u'1:안 켜짐',
      u'2:Hand로만 동작함',      
      u'3:증상 직접 입력' ,
      u'이전 메뉴'    
]
_HandpieceengineSymptomMultiChoiceList = [
      u'1:안 켜짐',
      u'2:Hand로만 동작함'      
]
_HandpieceengineSymptomJoinString = u'\n'.join(_HandpieceengineSymptomMultiChoiceList)


_AirinletSymptomKeyList = [
      u'1:흡입 안 됨',
      u'2:증상 직접 입력',
      u'이전 메뉴'    
]
_AirinletSymptomMultiChoiceList = [
      u'1:흡입 안 됨'
]
_AirinletSymptomJoinString = u'\n'.join(_AirinletSymptomMultiChoiceList)

_AiroutletSymptomKeyList = [
      u'1:공기 방출 안 됨',
      u'2:방출 단계 조절 안 됨',
      u'3:공기 방출구 빠짐' , 
      u'4:증상 직접 입력',
      u'이전 메뉴'    
]
_AiroutletSymptomMultiChoiceList = [
      u'1:공기 방출 안 됨',
      u'2:방출 단계 조절 안 됨',
      u'3:공기 방출구 빠짐' 
]
_AiroutletSymptomJoinString = u'\n'.join(_AiroutletSymptomMultiChoiceList)


_DollSymptomKeyList = [
      u'1:목 회전 안 됨' ,
      u'2:목 셕션 안 됨' , 
      u'3:상악 덴티폼 고정 나사 없음' , 
      u'4:하악 덴티폼 고정 나사 없음' ,       
      u'5:증상 직접 입력',
      u'이전 메뉴'        
]

_DollSymptomMultiChoiceList = [
      u'1:목 회전 안 됨' ,
      u'2:목 셕션 안 됨' ,
      u'3:상악 덴티폼 고정 나사 없음' , 
      u'4:하악 덴티폼 고정 나사 없음'        
]
_DollSymptomJoinString = u'\n'.join(_DollSymptomMultiChoiceList)

_3WaySymptomKeyList = [ 
      u'1:물 안 나옴' ,
      u'2:에어 안 나옴' ,
      u'3:증상 직접 입력',
      u'이전 메뉴'        
]

_3WaySymptomMultiChoiceList = [ 
      u'1:물 안 나옴' ,
      u'2:에어 안 나옴' 
]
_3WaySymptomJoinString = u'\n'.join(_3WaySymptomMultiChoiceList)


_HighspeedConnectorSymptomKeyList = [
      u'1:Connector 없음'  ,
      u'2:물 잠깐 나오다 안 나옴',             
      u'3:증상 직접 입력',
      u'이전 메뉴'        
]
_HighspeedConnectorSymptomMultiChoiceList = [
      u'1:Connector 없음'  ,
      u'2:물 잠깐 나오다 안 나옴',             
]
_HighspeedConnectorSymptomJoinString = u'\n'.join(_HighspeedConnectorSymptomMultiChoiceList)

_LowspeedConnectorSymptomKeyList = [
      u'1:Connector 없음' ,
      u'2:증상 직접 입력',
      u'이전 메뉴'        
]

_LowspeedConnectorSymptomMultiChoiceList = [
      u'1:Connector 없음' 
]
_LowspeedConnectorSymptomJoinString = u'\n'.join(_LowspeedConnectorSymptomMultiChoiceList)

_HighspeedHandpieceSymptomKeyList = [
      u'1:회전 안 됨' , 
      u'2:물 안 나옴',             
      u'3:증상 직접 입력', 
      u'이전 메뉴'       
]
_HighspeedHandpieceSymptomMultiChoiceList = [
      u'1:회전 안 됨' , 
      u'2:물 안 나옴'             
]
_HighspeedHandpieceSymptomJoinString = u'\n'.join(_HighspeedHandpieceSymptomMultiChoiceList)

_LowspeedHandpieceSymptomList = [
      u'1:회전 안 됨' , 
      u'2:증상 직접 입력', 
      u'이전 메뉴'       
]
_LowspeedHandpieceSymptomMultiChoiceList = [
      u'1:회전 안 됨' 
]
_LowspeedHandpieceSymptomJoinString = u'\n'.join(_LowspeedHandpieceSymptomMultiChoiceList)


_ComnetworkSymptomKeyList = [
      u'1:오프라인으로 나옴' ,
      u'2:IP 충돌이라고 나옴' , 
      u'3:증상 직접 입력', 
      u'이전 메뉴'       
]
_ComnetworkSymptomMultiChoiceList = [
      u'1:오프라인으로 나옴' ,
      u'2:IP 충돌이라고 나옴' , 
    ]   
_ComnetworkSymptomJoinString = u'\n'.join(_ComnetworkSymptomMultiChoiceList)

_CombodySymptomKeyList = [
      u'1:전원이 안 켜짐'  ,
      u'2:부팅이 안 됨' ,
      u'3:USB 인식을 못 함' ,
      u'이전 메뉴'       
]
_CombodySymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  ,
      u'2:부팅이 안 됨' ,
      u'3:USB 인식을 못 함'
]
_CombodySymptomJoinString = u'\n'.join(_CombodySymptomMultiChoiceList)


_Table1ButtonList = [
    u'Sand Blast 6', u'Sand Blast 5' , u'Vacuum Mixer 111',u'Sand Blast 4' , 
    u'Trimmer 10',   u'Sand Blast 3',  u'Sand Blast 7' ,
    u'직접 입력',      u'이전 메뉴'              
]
_Table1SNList = [
    u'1-13304400-000006', u'1-13304400-000005', u'1-12609100-000111', u'1-13304400-000004',
    u'1-13303200-000010', u'1-13304400-000003', u'1-13304400-000007'
]
_Table2ButtonList = [
    u'Trimmer 8'   , u'Trimmer 9'  ,   u'Trimmer 14'  ,
    u'직접 입력'     ,  u'이전 메뉴'              
]
_Table2SNList = [
    u'1-13301300-000008', u'1-13301300-000009', u'1-13303200-000014'
]
_Table3ButtonList = [
    u'Trimmer 10'   , u'Trimmer 17'  ,   u'Trimmer 12'  ,
    u'직접 입력'     ,  u'이전 메뉴'              
]
_Table3SNList = [
    u'1-13301300-000010', u'1-13303200-000017', u'1-13303200-000012'
]

_Table4ButtonList = [
    u'Trimmer 15'   , u'Trimmer 12'  ,   u'Trimmer 11'  ,
    u'직접 입력'     ,  u'이전 메뉴'              
]
_Table4SNList = [
    u'1-13303200-000015', u'1-13301300-000012', u'1-13303200-000011'
]

_Table5ButtonList = [
    u'캐스팅머신 12',     u'전기로 69',        u'전기로 68',
    u'전기로 67',       u'전기로 66',
    u'직접 입력'     ,   u'이전 메뉴'              
]
_Table5SNList = [
    u'1-13303000-000012', u'1-10500500-000069', u'1-10500500-000068',
    u'1-10500500-000067', u'1-10500500-000066'
]
_Table6ButtonList = [
    u'캐스팅머신 13',     u'전기로 65',        u'전기로 64',
    u'직접 입력'     ,   u'이전 메뉴'              
]
_Table6SNList = [
    u'1-13303000-000013', u'1-10500500-000065', u'1-10500500-000064'
]

_Table7ButtonList = [
    u'캐스팅머신 14',     u'전기로 63',        u'전기로 62',
    u'직접 입력'     ,   u'이전 메뉴'              
]
_Table7SNList = [
    u'1-13303000-000014', u'1-10500500-000063', u'1-10500500-000062'
]

_Table8ButtonList = [
    u'온성기 11',        u'온성기 13',       u'스팀크리너 78',
    u'스팀크리너 79',     u'온성기 12', 
    u'직접 입력'     ,   u'이전 메뉴'              
]
_Table8SNList = [
    u'1-13302500-000011', u'1-13302500-000013', u'1-12627800-000078',
    u'1-12627800-000079', u'1-13302500-000012'
]
_Table9ButtonList = [
    u'분배기 1',         u'분배기 2',        u'스팀크리너 42',
    u'Vacuum Mixer 109',u'Vacuum Mixer 102', 
    u'직접 입력'     ,   u'이전 메뉴'         
]
_Table9SNList = [
    u'2-12611300-000001', u'2-12611300-000002', u'1-12627800-000042',
    u'1-12609100-000109', u'1-12609100-000102'
]
_Table10ButtonList = [
    u'Trimmer 9'  ,    u'Vacuum Mixer 108', u'분배기 4', u'분배기 3',
    u'직접 입력'     ,   u'이전 메뉴'              
]
_Table10SNList = [
    u'1-13303200-000009', u'1-12609100-000108', u'2-12611300-000004', u'2-12611300-000003'
]

_Table11ButtonList = [
    u'Trimmer 16'  ,    u'Vacuum Mixer 110', u'스팀크리너 41',
    u'직접 입력'     ,   u'이전 메뉴'                  
]
_Table11SNList = [
    u'1-13303200-000016', u'1-12609100-000110', u'1-12627800-000041'  
]

_Table12ButtonList = [
    u'Poll Cleaner 48',u'Poll Cleaner 51',u'Poll Cleaner 52',u'Poll Cleaner 53',
    u'직접 입력'     ,   u'이전 메뉴'                  
]
_Table12SNList = [
    u'1-12627800-000048', u'1-12627800-000051', u'1-12627800-000052', u'1-12627800-000053'
]

_TableButtonListList = [
    _Table1ButtonList,  _Table2ButtonList, _Table3ButtonList, _Table4ButtonList, 
    _Table5ButtonList,  _Table6ButtonList, _Table7ButtonList, _Table8ButtonList,
    _Table9ButtonList,  _Table10ButtonList,_Table11ButtonList,_Table12ButtonList         
]

_TableSNListList = [
    _Table1SNList,  _Table2SNList, _Table3SNList, _Table4SNList, 
    _Table5SNList,  _Table6SNList, _Table7SNList, _Table8SNList,
    _Table9SNList,  _Table10SNList,_Table11SNList,_Table12SNList         
]

_SandblastSymptomKeyList  = [
      u'1:전원이 안 켜짐'  ,
      u'2:바람은 나오나 모래가 안 나옴', 
      u'3:증상 직접 입력', 
      u'이전 메뉴'       
]

_SandblastSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  ,
      u'2:바람은 나오나 모래가 안 나옴' 
]
_SandblastSymptomJoinString = u'\n'.join(_SandblastSymptomMultiChoiceList)
_VacuummixerSymptomKeyList = [
      u'1:전원이 안 켜짐'  ,
      u'2:비커가 회전이 안 됨', 
      u'3:증상 직접 입력', 
      u'이전 메뉴'       
]
_VacuummixerSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  ,
      u'2:비커가 회전이 안 됨' 
]
_VacuummixerSymptomJoinString = u'\n'.join(_VacuummixerSymptomMultiChoiceList)
_TrimmerSymptomKeyList = [
      u'1:전원이 안 켜짐'  ,
      u'2:날이 회전이 안 됨' , 
      u'3:증상 직접 입력', 
      u'이전 메뉴'       
]

_TrimmerSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  ,
      u'2:날이 회전이 안 됨' 
]
_TrimmerSymptomJoinString = u'\n'.join(_TrimmerSymptomMultiChoiceList)
_CastingmachineSymptomKeyList = [
      u'1:전원이 안 켜짐'  , 
      u'2:증상 직접 입력', 
      u'이전 메뉴'       
]

_CastingmachineSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  
]
_CastingmachineSymptomJoinString = u'\n'.join(_CastingmachineSymptomMultiChoiceList)
_ElectricfurnaceSymptomKeyList = [
      u'1:전원이 안 켜짐'  , 
      u'2:증상 직접 입력', 
      u'이전 메뉴'       
]

_ElectricfurnaceSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  
]
_ElectricfurnaceSymptomJoinString = u'\n'.join(_ElectricfurnaceSymptomMultiChoiceList)
_CuringwaterbathSymptomKeyList = [
      u'1:전원이 안 켜짐'  , 
      u'2:증상 직접 입력', 
      u'이전 메뉴'       
]

_CuringwaterbathSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  
]
_CuringwaterbathSymptomJoinString = u'\n'.join(_CuringwaterbathSymptomMultiChoiceList)
_SteamcleanerSymptomKeyList = [
      u'1:전원이 안 켜짐' , 
      u'2:증상 직접 입력', 
      u'이전 메뉴'        
]

_SteamcleanerSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  
]
_SteamcleanerSymptomJoinString = u'\n'.join(_SteamcleanerSymptomMultiChoiceList)
_DispenserSymptomKeyList = [
      u'1:전원이 안 켜짐'  , 
      u'2:증상 직접 입력', 
      u'이전 메뉴'        
]

_DispenserSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  
]
_DispenserSymptomJoinString = u'\n'.join(_DispenserSymptomMultiChoiceList)
_PollcleanerSymptomKeyList = [
      u'1:전원이 안 켜짐'  , 
      u'2:증상 직접 입력', 
      u'이전 메뉴'        
]

_PollcleanerSymptomMultiChoiceList = [
      u'1:전원이 안 켜짐'  
]
_PollcleanerSymptomJoinString = u'\n'.join(_PollcleanerSymptomMultiChoiceList)


def  _nx_Child_in(stage_num , score) :
    if score == 0 :
        return stage_num
    else :
        return _nx_Child_in(stage_num * 0x10 +1 , score-1)
def  nx_Child_in( stage_num , score ) :
    num = _nx_Child_in(stage_num, score)
    return num
def nx_Child_Sibling_in(stage_num , child_score, sibling_score) :
    num = nx_Child_in(stage_num, child_score) + sibling_score
    return num

initial_State          = 0x1
first_4work_State      = 0x111111
first_3com_State       = 0x111115
first_3work_State      = 0x111113
first_3handpiece_State = 0x111114
first_4eng_State       = 0x111112
first_Independent_IDInsert_State = 0x141

StateMultiChoiceList = {
                nx_Child_in(first_4work_State,3): _LightSymptomMultiChoiceList  ,            nx_Child_Sibling_in(first_4work_State,3,1): _MonitorSymptomMultiChoiceList ,     
                nx_Child_Sibling_in(first_4work_State,3,2):_GastorchSymptomMultiChoiceList , nx_Child_Sibling_in(first_4work_State,3,3): _HandpieceengineSymptomMultiChoiceList, 
                nx_Child_Sibling_in(first_4work_State,3,4):_AirinletSymptomMultiChoiceList,  nx_Child_Sibling_in(first_4work_State,3,5): _AiroutletSymptomMultiChoiceList ,  
                nx_Child_in(first_3work_State,3): _DollSymptomMultiChoiceList ,              nx_Child_Sibling_in(first_3work_State,3,1):_MonitorSymptomMultiChoiceList , 
                nx_Child_Sibling_in(first_3work_State,3,2):_LightSymptomMultiChoiceList ,    nx_Child_Sibling_in(first_3work_State,3,3):_3WaySymptomMultiChoiceList,      
                nx_Child_Sibling_in(first_3work_State,3,4):_HighspeedConnectorSymptomMultiChoiceList, nx_Child_Sibling_in(first_3work_State,3,5):_LowspeedConnectorSymptomMultiChoiceList ,
                nx_Child_in(first_3handpiece_State,3): _HighspeedHandpieceSymptomMultiChoiceList ,    nx_Child_Sibling_in(first_3handpiece_State,3,1): _LowspeedHandpieceSymptomMultiChoiceList   ,        
                nx_Child_in(first_3com_State,3): _MonitorSymptomMultiChoiceList ,  nx_Child_Sibling_in(first_3com_State,3,1):_CombodySymptomMultiChoiceList ,  
                nx_Child_Sibling_in(first_3com_State,3,2):_ComnetworkSymptomMultiChoiceList
}

StateButtonList = { initial_State: _State0KeyList, 
                    nx_Child_Sibling_in(initial_State,1,1): _State_ItemListCheckTop_List ,                                            
                  nx_Child_Sibling_in(initial_State,1,3): _State4KeyList ,
                  nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,1),1,1) : _YesorNoKeyList ,
                  nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,1) : _YesorNoKeyList ,   
                  nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,2) : _InputModeList, 
                  nx_Child_in(nx_Child_Sibling_in(initial_State,1,1),2): _YesorNoKeyList  ,
                 nx_Child_in(first_Independent_IDInsert_State,2): _GradeKeyList ,                 
                 nx_Child_in(first_Independent_IDInsert_State,3): _YesorNoKeyList  ,   
                 nx_Child_in(initial_State,3): _GradeKeyList ,  
                 ##nx_Child_Sibling_in(initial_State,4,2):  _State13KeyList,            
                 nx_Child_in(initial_State,4): _State1KeyList ,                      
                 nx_Child_in(first_4work_State,1): _State111KeyList,          nx_Child_in(first_3com_State,1): _State141KeyList,  
                 nx_Child_in(first_3work_State,1): _State1311KeyList ,       nx_Child_in(first_3handpiece_State,1) : _State1321KeyList, 
                 nx_Child_in(first_4work_State,3): _LightSymptomKeyList  ,         nx_Child_Sibling_in(first_4work_State,3,1): _MonitorSymptomKeyList ,     
                 nx_Child_Sibling_in(first_4work_State,3,2) : _GastorchSymptomKeyList , nx_Child_Sibling_in(first_4work_State,3,3): _HandpieceengineSymptomKeyList ,
                 nx_Child_Sibling_in(first_4work_State,3,4): _AirinletSymptomKeyList ,    nx_Child_Sibling_in(first_4work_State,3,5): _AiroutletSymptomKeyList ,
                 nx_Child_in(first_3work_State,3): _DollSymptomKeyList ,     nx_Child_Sibling_in(first_3work_State,3,1):_MonitorSymptomKeyList , 
                 nx_Child_Sibling_in(first_3work_State,3,2):_LightSymptomKeyList ,  nx_Child_Sibling_in(first_3work_State,3,3):_3WaySymptomKeyList ,      
                 nx_Child_Sibling_in(first_3work_State,3,4): _HighspeedConnectorSymptomKeyList   ,     nx_Child_Sibling_in(first_3work_State,3,5):_LowspeedConnectorSymptomKeyList ,
                 nx_Child_in(first_3handpiece_State,3): _HighspeedHandpieceSymptomKeyList,    nx_Child_Sibling_in(first_3handpiece_State,3,1): _LowspeedHandpieceSymptomList,                
                 nx_Child_in(first_3com_State,3): _MonitorSymptomKeyList ,  nx_Child_Sibling_in(first_3com_State,3,1):_CombodySymptomKeyList,  
                 nx_Child_Sibling_in(first_3com_State,3,2):_ComnetworkSymptomKeyList ,
                 nx_Child_in(first_4work_State,5): _YesorNoKeyListv2,    nx_Child_in(first_3com_State,5): _YesorNoKeyListv2,   #1111411111: _YesorNoKeyList,
                 nx_Child_in(first_3work_State,5): _YesorNoKeyListv2,    nx_Child_in(first_3handpiece_State,5): _YesorNoKeyListv2                      
}

StatePhotoList = {  
                    first_4work_State:   { "url": u'static/images/4work_seats.png' , "width": 493, "height": 682 } ,
                    nx_Child_in(first_4work_State,1):  {"url": u'static/images/4work_oneseat.png' ,"width": 200, "height": 283 } ,
                    first_3work_State:   { "url": u'static/images/4work_seats.png'  ,"width": 493 ,"height": 682 }, 
                    nx_Child_in(first_3work_State,1): {"url": u'static/images/3work_oneseat.png' ,"width": 200, "height": 283 } ,
                    first_3handpiece_State:  {"url": u'static/images/3work_case.png' ,"width": 230,"height": 218}, 
                    nx_Child_in(first_3handpiece_State,1): {"url": u'static/images/3work_caseopen.png' ,"width": 200,"height": 283 } ,
                    first_3com_State:   {"url": u'static/images/3com_seats.png' ,   "width": 542,"height": 611 }, 
                    nx_Child_in(first_3com_State,1):  {"url": u'static/images/3com_oneseat.png' ,"width": 363,"height": 616 }
#                    141:  {"url": request.url_root+u'static/images/3com_oneseat.png' ,"width": 363,"height": 616 }
}

SelectString = u' 선택하셨습니다.'
InsertedString = u' 입력하셨습니다.'
SubmitString = u'접수되었습니다.'
CancelString = u'취소되었습니다'
UnderConstructionString =u'-Under Construction-'
UnInsertedString = u'필수 항목인 학번(혹은 사번)이 입력되지 않았습니다.'
SameButtonString = u'accessing cloud...'
ExplainSymptomInsertionString = u'----------------------------------------\n'
ExplainSymptomInsertionString +=u'ex) (복수 입력 가능)\n\t\t 2\n\t\t 1 2\n\t\t 물이 샘\n\t\t 1 2 물이 샘'
#ExplainSymptomInsertionString +=u'ex) (복수 입력 가능)\n\t\t 2\n\t\t 1,2\n\t\t 물이 샘\n\t\t 1,2,물이 샘'
#ExplainSymptomInsertionString +=u'ex) 2\n\t\t 1,2\n\t\t 물이 샘\n\t\t 1,2,물이 샘'
#ExplainSymptomInsertionString +=u'단/복수 입력이 가능합니다.\nex) 2 (객관식 단수)\n\t\t물이 샘 (주관식 단수)\n\t\t1,2 (객관식 복수)\n\t\t1,2,물이 샘 (혼합 복수)'

AskLocationString = u'위치가 어디신가요?'
AskSeatNumberString = u'자리가 어디신가요?\n0:이전 메뉴'
AskTableNumberString = u'테이블이 어디신가요?\n0:이전 메뉴'
AskPartString = u'어떤 부분이 문제인가요?'
AskDeviceString = u'어떤 기계가 문제인가요?'
AskSymptomString = u'어떤 증상인가요?'
AskMultiSymptomString = u'어떤 증상인가요?\n\n0:이전 메뉴로 돌아가기'
InsertIDString = u'학번(혹은 사번)을 입력해주세요 ex)2011740011\n0:이전 메뉴'
InsertNameString = u'이름(혹은 별명)을 입력해주세요 ex)오승환, 강정호a, HyunsooKim\n0:이전 메뉴'
InsertGradeString = u'학년을 입력해 주세요'
ReInsertString = u'다시 입력해 주세요'
InsertYesNoString = u'입력하시겠습니까?'
LastYesNoString = u'최종 접수하시겠습니까'
DirectInsertSymptomString = u'직접 증상을 입력해주세요\n0:이전 메뉴'
DirectInsertPartString = u'고장난 부분과 증상을 한꺼번에 입력해주세요\n0:이전 메뉴'
DirectInsertDeviceString = u'고장난 기계와 증상을 한꺼번에 입력해주세요\n0:이전 메뉴'

InsertValidNumberString = u'범위 내의 숫자를 입력해주세요'
InsertNumberString = u'숫자를 입력해주세요'
InsertCaseNumberString = u'Case의 번호를 입력해주세요\n0:이전 메뉴'
InsertItemNumberString = u'삭제할 Item의 번호를 모두 입력해주세요\n\n예시)\n2: 2번 입력\n1 3: 1과 3번 입력\n0: 이전 메뉴'


AskSeatHandpieceString = u'실습실 자리 문제인가요? 핸드피스 문제인가요?'
AskDeletionString = u'삭제하시겠습니까?'
AskPasswordString = u'Password를 입력해주세요'
AskMovementString = u'이동하실 메뉴를 선택해주세요\n ex)1:초기메뉴'
AskInputModeString = u'어떤 입력 방식으로  변경하시겠습니까?'

fromStateMessageList = {  initial_State:SelectString+u'\n' ,
                          nx_Child_in(initial_State,1):SelectString+u'\n' ,
                          nx_Child_Sibling_in(initial_State,1,1):SelectString+u'\n', 
                          first_Independent_IDInsert_State:SelectString+u'\n' ,        
                          nx_Child_Sibling_in(initial_State,1,3):SelectString+u'\n' ,
                          nx_Child_in(nx_Child_Sibling_in(initial_State,1,1),1):SelectString+ u'\n',
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,1),1,1):SelectString+ u'\n',
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,1):SelectString+u'\n' ,            
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,2):SelectString+u'\n' ,         
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,4):SelectString+u'\n' , 
                          nx_Child_in(nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,4),1):SelectString+u'\n' ,         
                          nx_Child_in(initial_State,2):SelectString+u'\n' ,          
                          nx_Child_in(first_Independent_IDInsert_State,1):SelectString+ u'\n' ,                                           
                          nx_Child_in(initial_State,3):SelectString+u'\n' ,
                          nx_Child_in(initial_State,4):SelectString+u'\n' ,
                          nx_Child_in(nx_Child_Sibling_in(initial_State,1,1),2):SelectString+ u'\n',
                          nx_Child_in(first_Independent_IDInsert_State,2):SelectString+ u'\n',
                          nx_Child_in(first_Independent_IDInsert_State,3):SelectString+ u'\n',
                          first_4work_State:InsertedString+u'\n' ,    first_3com_State:SelectString+u'\n' ,          
                          first_3work_State:SelectString+u'\n' ,       first_3handpiece_State:SelectString+u'\n' ,                                                 
                          nx_Child_in(first_4work_State,1):InsertedString+u'\n' ,   nx_Child_in(first_3com_State,1):SelectString+u'\n' ,         
                          nx_Child_in(first_3work_State,1):SelectString+u'\n' ,     nx_Child_in(first_3handpiece_State,1):SelectString+u'\n' ,
                          nx_Child_in(first_4work_State,2):SelectString+u'\n' ,     nx_Child_in(first_3com_State,2):SelectString+u'\n' ,        
                          nx_Child_in(first_3work_State,2):SelectString+u'\n' ,     nx_Child_in(first_3handpiece_State,2):SelectString+u'\n' ,
                          nx_Child_in(first_4work_State,3):SelectString+u'\n' ,     nx_Child_in(first_3com_State,3):SelectString+u'\n' ,       
                          nx_Child_in(first_3work_State,3):SelectString+u'\n' ,     nx_Child_in(first_3handpiece_State,3):SelectString+u'\n' ,
                          nx_Child_Sibling_in(first_4work_State,3,1):SelectString+u'\n' ,   nx_Child_Sibling_in(first_3com_State,3,1):SelectString+u'\n' ,       
                          nx_Child_Sibling_in(first_3work_State,3,1):SelectString+u'\n' ,   nx_Child_Sibling_in(first_3handpiece_State,3,1):SelectString+u'\n' ,
                          nx_Child_Sibling_in(first_4work_State,3,2):SelectString+u'\n' ,   nx_Child_Sibling_in(first_3com_State,3,2):SelectString+u'\n' ,       
                          nx_Child_Sibling_in(first_3work_State,3,2):SelectString+u'\n' ,                           
                          nx_Child_Sibling_in(first_4work_State,3,3):SelectString+u'\n' ,   nx_Child_Sibling_in(first_3work_State,3,3):SelectString+u'\n' ,  
                          nx_Child_Sibling_in(first_4work_State,3,4):SelectString+u'\n' ,   nx_Child_Sibling_in(first_3work_State,3,4):SelectString+u'\n' ,
                          nx_Child_Sibling_in(first_4work_State,3,5):SelectString+u'\n' ,   nx_Child_Sibling_in(first_3work_State,3,5):SelectString+u'\n' ,
                          nx_Child_in(first_4work_State,4):SelectString+u'\n' ,  nx_Child_in(first_3com_State,4):SelectString+u'\n' ,      
                          nx_Child_in(first_3work_State,4):SelectString+u'\n' ,   nx_Child_in(first_3handpiece_State,4):SelectString+u'\n' ,     
                          nx_Child_in(first_4work_State,5):SelectString+SubmitString+u'\n' ,nx_Child_in(first_3com_State,5):SelectString+SubmitString+u'\n' ,
                          nx_Child_in(first_3work_State,5):SelectString+SubmitString+u'\n' ,nx_Child_in(first_3handpiece_State,5):SelectString+SubmitString+u'\n' 
}

toStateMessageList = {    initial_State:u'',
                          nx_Child_in(initial_State,1):InsertIDString,
                          first_Independent_IDInsert_State:InsertIDString,
                          nx_Child_Sibling_in(initial_State,1,1):u'', 
                          nx_Child_Sibling_in(initial_State,1,3):u'',
                          nx_Child_in(initial_State,2):InsertNameString,
                          nx_Child_in(nx_Child_Sibling_in(initial_State,1,1),1):InsertItemNumberString,
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,1),1,1):AskDeletionString,                          
                          nx_Child_in(nx_Child_Sibling_in(initial_State,1,1),2):AskDeletionString,
                          nx_Child_in(first_Independent_IDInsert_State,1):InsertNameString,
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,1):AskDeletionString,                                 
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,2): AskInputModeString,               
                          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,4):AskPasswordString,
                          nx_Child_in(nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,4),1):AskMovementString,
                          nx_Child_in(initial_State,3):InsertGradeString,
                          nx_Child_in(first_Independent_IDInsert_State,2):InsertGradeString,
                          #nx_Child_Sibling_in(initial_State,4,2):AskSeatHandpieceString,                                                    
                          nx_Child_in(initial_State,4):AskLocationString,                                                 
                          first_4work_State:AskSeatNumberString,       first_3com_State:AskSeatNumberString,                 
                          first_3work_State:AskSeatNumberString,       first_3handpiece_State:InsertCaseNumberString,                          
                          nx_Child_in(first_4work_State,1):AskPartString,            nx_Child_in(first_3com_State,1):AskPartString,                      
                          nx_Child_in(first_3work_State,1):AskPartString,            nx_Child_in(first_3handpiece_State,1):AskPartString,       
                          nx_Child_in(first_4work_State,2):DirectInsertPartString,  nx_Child_in(first_3com_State,2):DirectInsertPartString,
                          nx_Child_in(first_3work_State,2):DirectInsertPartString,  nx_Child_in(first_3handpiece_State,2):DirectInsertPartString,
                          nx_Child_in(first_3com_State,3):AskSymptomString ,                  
                          nx_Child_Sibling_in(first_3com_State,3,1):AskSymptomString ,                                            
                          nx_Child_Sibling_in(first_3com_State,3,2):AskSymptomString ,
                          nx_Child_in(first_4work_State,3):AskSymptomString ,       
                          nx_Child_Sibling_in(first_4work_State,3,1):AskSymptomString ,                                                   
                          nx_Child_Sibling_in(first_4work_State,3,2):AskSymptomString ,
                          nx_Child_Sibling_in(first_4work_State,3,3):AskSymptomString ,
                          nx_Child_Sibling_in(first_4work_State,3,4):AskSymptomString ,
                          nx_Child_Sibling_in(first_4work_State,3,5):AskSymptomString ,
                          nx_Child_in(first_3work_State,3):AskSymptomString,       
                          nx_Child_Sibling_in(first_3work_State,3,1):AskSymptomString,    
                          nx_Child_Sibling_in(first_3work_State,3,2):AskSymptomString,
                          nx_Child_Sibling_in(first_3work_State,3,3):AskSymptomString,
                          nx_Child_Sibling_in(first_3work_State,3,4):AskSymptomString,
                          nx_Child_Sibling_in(first_3work_State,3,5):AskSymptomString ,        
                          nx_Child_in(first_3handpiece_State,3):AskSymptomString,   
                          nx_Child_Sibling_in(first_3handpiece_State,3,1):AskSymptomString, 
                          nx_Child_in(first_4work_State,4):DirectInsertSymptomString,  nx_Child_in(first_3com_State,4):DirectInsertSymptomString,       
                          nx_Child_in(first_3work_State,4):DirectInsertSymptomString,  nx_Child_in(first_3handpiece_State,4):DirectInsertSymptomString
}

toSymptomStateMessageListsList = {
                nx_Child_in(first_4work_State,3): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_LightSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,         
                nx_Child_Sibling_in(first_4work_State,3,1): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_MonitorSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] , 
                nx_Child_Sibling_in(first_4work_State,3,2): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_GastorchSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,      
                nx_Child_Sibling_in(first_4work_State,3,3): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_HandpieceengineSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] , 
                nx_Child_Sibling_in(first_4work_State,3,4): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_AirinletSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,     
                nx_Child_Sibling_in(first_4work_State,3,5): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_AiroutletSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,   
                nx_Child_in(first_3work_State,3):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_DollSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString],       
                nx_Child_Sibling_in(first_3work_State,3,1):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_MonitorSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString],    
                nx_Child_Sibling_in(first_3work_State,3,2):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_LightSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString],
                nx_Child_Sibling_in(first_3work_State,3,3):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_3WaySymptomJoinString+u'\n\n'+ExplainSymptomInsertionString],
                nx_Child_Sibling_in(first_3work_State,3,4):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_HighspeedConnectorSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString],
                nx_Child_Sibling_in(first_3work_State,3,5):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_LowspeedConnectorSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,       
                nx_Child_in(first_3handpiece_State,3):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_HighspeedHandpieceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,    
                nx_Child_Sibling_in(first_3handpiece_State,3,1):[AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_LowspeedHandpieceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,   
                nx_Child_in(first_3com_State,3): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_MonitorSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,                  
                nx_Child_Sibling_in(first_3com_State,3,1): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_CombodySymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] ,                                            
                nx_Child_Sibling_in(first_3com_State,3,2): [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ComnetworkSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString] 

}

push_StateList = {
                  nx_Child_in(first_4work_State,2):True, 
                  nx_Child_in(first_4work_State,3):True, nx_Child_Sibling_in(first_4work_State,3,1):True, 
                  nx_Child_Sibling_in(first_4work_State,3,2):True, nx_Child_Sibling_in(first_4work_State,3,3):True, 
                  nx_Child_Sibling_in(first_4work_State,3,4):True, nx_Child_Sibling_in(first_4work_State,3,5):True, 
                  nx_Child_in(first_4work_State,4):True,
                  nx_Child_in(first_3com_State,2):True, 
                  nx_Child_in(first_3com_State,3):True, nx_Child_Sibling_in(first_3com_State,3,1):True, 
                  nx_Child_Sibling_in(first_3com_State,3,2):True,                                                    
                  nx_Child_in(first_3com_State,4):True,
                  nx_Child_in(first_3work_State,2):True,
                  nx_Child_in(first_3work_State,3):True,nx_Child_Sibling_in(first_3work_State,3,1):True,
                  nx_Child_Sibling_in(first_3work_State,3,2):True,nx_Child_Sibling_in(first_3work_State,3,3):True,
                  nx_Child_Sibling_in(first_3work_State,3,4):True,nx_Child_Sibling_in(first_3work_State,3,5):True,
                  nx_Child_in(first_3work_State,4):True,                
                  nx_Child_in(first_3handpiece_State,2):True,
                  nx_Child_in(first_3handpiece_State,3):True,nx_Child_Sibling_in(first_3handpiece_State,3,1):True,                                                                    
                  nx_Child_in(first_3handpiece_State,4):True
}

pop_pushedStateList = {
                       nx_Child_in(first_4work_State,4):True , nx_Child_in(first_3com_State,4):True , 
                       nx_Child_in(first_3work_State,4):True ,  nx_Child_in(first_3handpiece_State,4):True     
}   


state = { initial_State:0x1 , 
          nx_Child_in(initial_State,1):0x11,
          nx_Child_Sibling_in(initial_State,1,1):0x12,                                            
          nx_Child_Sibling_in(initial_State,1,3):0x14 , 
          nx_Child_in(initial_State,2):0x111,
          nx_Child_in(nx_Child_Sibling_in(initial_State,1,1),1):0x121,      
          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,1),1,1):0x122, 
          first_Independent_IDInsert_State:0x141,                   
          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,1):0x142,              
          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,2):0x143,                       
          nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,4):0x145,     
          nx_Child_in(nx_Child_Sibling_in(initial_State,1,1),2):0x1211,
          nx_Child_in(first_Independent_IDInsert_State,1):0x1411,
          nx_Child_in(first_Independent_IDInsert_State,2):0x14111 ,   
          nx_Child_in(first_Independent_IDInsert_State,3):0x141111 ,
          nx_Child_in(nx_Child_Sibling_in(nx_Child_Sibling_in(initial_State,1,3),1,4),1):0x1451,
          nx_Child_in(initial_State,3):0x1111,                                
          ##nx_Child_Sibling_in(initial_State,4,2):0x11113, 
          nx_Child_in(initial_State,4):0x11111,                                 
          first_4work_State:0x111111 ,                                first_3com_State:0x111115,                                         
          first_3work_State:0x111113,                                 first_3handpiece_State:0x111114,                            
          nx_Child_in(first_4work_State,1):0x1111111 ,                nx_Child_in(first_3com_State,1):0x1111151,       
          nx_Child_in(first_3work_State,1):0x1111131,                 nx_Child_in(first_3handpiece_State,1):0x1111141,  
          nx_Child_in(first_4work_State,2):0x11111111 ,               nx_Child_in(first_3com_State,2):0x11111511,   
          nx_Child_in(first_3work_State,2):0x11111311,                nx_Child_in(first_3handpiece_State,2):0x11111411,                                    
          nx_Child_in(first_4work_State,3):0x111111111 ,              nx_Child_in(first_3com_State,3):0x111115111,               
          nx_Child_in(first_3work_State,3):0x111113111  ,             nx_Child_in(first_3handpiece_State,3):0x111114111  ,  
          nx_Child_Sibling_in(first_4work_State,3,1):0x111111112,     nx_Child_Sibling_in(first_3com_State,3,1):0x111115112,
          nx_Child_Sibling_in(first_3work_State,3,1):0x111113112 ,    nx_Child_Sibling_in(first_3handpiece_State,3,1):0x111114112 ,
          nx_Child_Sibling_in(first_4work_State,3,2):0x111111113,     nx_Child_Sibling_in(first_3com_State,3,2):0x111115113,  
          nx_Child_Sibling_in(first_3work_State,3,2):0x111113113 ,
          nx_Child_Sibling_in(first_4work_State,3,3):0x111111114 ,    nx_Child_Sibling_in(first_3work_State,3,3):0x111113114 ,        
          nx_Child_Sibling_in(first_4work_State,3,4):0x111111115,     nx_Child_Sibling_in(first_3work_State,3,4):0x111113115 ,
          nx_Child_Sibling_in(first_4work_State,3,5):0x111111116,     nx_Child_Sibling_in(first_3work_State,3,5):0x111113116 ,                          
          nx_Child_in(first_4work_State,4):0x1111111111,              nx_Child_in(first_3com_State,4):0x1111151111,
          nx_Child_in(first_3work_State,4):0x1111131111 ,             nx_Child_in(first_3handpiece_State,4):0x1111141111 ,                                                                                                   
          nx_Child_in(first_4work_State,5):0x11111111111,             nx_Child_in(first_3com_State,5):0x11111511111,    
          nx_Child_in(first_3work_State,5):0x11111311111,             nx_Child_in(first_3handpiece_State,5):0x11111411111,              
          0x1111132113:0x1111132113  #??           
}

_4EngSymptomStateList = []
_4EngDevSNList = {}

len_4eng_tables = 0xc
def generate4EngStatesInformation() :

    len_tables = len_4eng_tables
    #len_devices_per_table = [7,3,3,3, 5,3,3,5, 5,4,3,4 ]
    len_devices_per_table = []
    for i in range(len_tables)  :
        len_devices_per_table.append( len('_Table'+ str(i+1) +'ButtonList') - 2 )    

    state[first_4eng_State] = first_4eng_State   # which table?
    for i in range(len_tables)  :
        state[ nx_Child_in(first_4eng_State,1)+i ]   =  nx_Child_in(first_4eng_State,1)+i  #which device?
        _current_State = nx_Child_in(first_4eng_State,1)+i
        state[ nx_Child_in(_current_State,1) ] = nx_Child_in(_current_State,1)      # insert device directly
        _current_State = nx_Child_in(_current_State,1)
        for j in range(len(len_devices_per_table)) : 
            state[ nx_Child_in(_current_State,1)+j ] = nx_Child_in(_current_State,1)+j      #which Symptom?
            _4EngSymptomStateList.append(nx_Child_in(_current_State,1)+j)
        _current_State = nx_Child_in(_current_State,1)
        state[ nx_Child_in(_current_State,1) ] = nx_Child_in(_current_State,1)              #insert symptom directly
        _current_State = nx_Child_in(_current_State,1)
        state[ nx_Child_in(_current_State,1) ] = nx_Child_in(_current_State,1)              #Y or Y+ or N?

    StatePhotoList[first_4eng_State] = {"url": u'static/images/4eng_tables.png' , "width": 626, "height": 660 }
    for i in range(len_tables)  :
        StatePhotoList[nx_Child_in(first_4eng_State,1)+i] = {"url": u'static/images/table'+str(i+1)+u'.jpg' , "width": 720, "height": 405 }
    StatePhotoList[nx_Child_in(first_4eng_State,1)+0]["height"] = 387

    for i in range(len_tables)  :
        StateButtonList[nx_Child_in(first_4eng_State,1)+i] = _TableButtonListList[i]
        StateButtonList[  nx_Child_in(nx_Child_in(first_4eng_State,1)+i ,4) ] = _YesorNoKeyListv2
    for i in range(len_tables)  :
        _4EngDevSNList[nx_Child_in(first_4eng_State,1)+i] = _TableSNListList[i]   

    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 0 , 2)
    StateMultiChoiceList[_current_State+0] =  _SandblastSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _SandblastSymptomMultiChoiceList  
    StateMultiChoiceList[_current_State+2] =  _VacuummixerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+3] =  _SandblastSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+4] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+5] =  _SandblastSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+6] =  _SandblastSymptomMultiChoiceList
    # len_devices_per_table[0]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 1 , 2)
    StateMultiChoiceList[_current_State+0] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+2] =  _TrimmerSymptomMultiChoiceList
    # len_devices_per_table[1]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 2 , 2)
    StateMultiChoiceList[_current_State+0] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+2] =  _TrimmerSymptomMultiChoiceList
    # len_devices_per_table[2]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 3 , 2)
    StateMultiChoiceList[_current_State+0] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+2] =  _TrimmerSymptomMultiChoiceList
    # len_devices_per_table[3]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 4 , 2)
    StateMultiChoiceList[_current_State+0] =  _CastingmachineSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _ElectricfurnaceSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _ElectricfurnaceSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+3] =  _ElectricfurnaceSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+4] =  _ElectricfurnaceSymptomMultiChoiceList    
    # len_devices_per_table[4]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 5 , 2)
    StateMultiChoiceList[_current_State+0] =  _CastingmachineSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _ElectricfurnaceSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _ElectricfurnaceSymptomMultiChoiceList    
    # len_devices_per_table[5]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 6 , 2)
    StateMultiChoiceList[_current_State+0] =  _CastingmachineSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _ElectricfurnaceSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _ElectricfurnaceSymptomMultiChoiceList    
    # len_devices_per_table[6]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 7 , 2)
    StateMultiChoiceList[_current_State+0] =  _CuringwaterbathSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _CuringwaterbathSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _SteamcleanerSymptomMultiChoiceList   
    StateMultiChoiceList[_current_State+3] =  _SteamcleanerSymptomMultiChoiceList   
    StateMultiChoiceList[_current_State+4] =  _CuringwaterbathSymptomMultiChoiceList
    # len_devices_per_table[7]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 8 , 2)
    StateMultiChoiceList[_current_State+0] =  _DispenserSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _DispenserSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _SteamcleanerSymptomMultiChoiceList   
    StateMultiChoiceList[_current_State+3] =  _VacuummixerSymptomMultiChoiceList   
    StateMultiChoiceList[_current_State+4] =  _VacuummixerSymptomMultiChoiceList
    # len_devices_per_table[8]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 9 , 2)
    StateMultiChoiceList[_current_State+0] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _VacuummixerSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _DispenserSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+3] =  _DispenserSymptomMultiChoiceList
    # len_devices_per_table[9]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 10 , 2)
    StateMultiChoiceList[_current_State+0] =  _TrimmerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _VacuummixerSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _SteamcleanerSymptomMultiChoiceList
    # len_devices_per_table[10]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 11 , 2)
    StateMultiChoiceList[_current_State+0] =  _PollcleanerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+1] =  _PollcleanerSymptomMultiChoiceList    
    StateMultiChoiceList[_current_State+2] =  _PollcleanerSymptomMultiChoiceList
    StateMultiChoiceList[_current_State+3] =  _PollcleanerSymptomMultiChoiceList
    # len_devices_per_table[11]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 0 , 2)
    StateButtonList[_current_State+0] =  _SandblastSymptomKeyList
    StateButtonList[_current_State+1] =  _SandblastSymptomKeyList  
    StateButtonList[_current_State+2] =  _VacuummixerSymptomKeyList
    StateButtonList[_current_State+3] =  _SandblastSymptomKeyList
    StateButtonList[_current_State+4] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+5] =  _SandblastSymptomKeyList
    StateButtonList[_current_State+6] =  _SandblastSymptomKeyList
    # len_devices_per_table[0]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 1 , 2)
    StateButtonList[_current_State+0] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+1] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+2] =  _TrimmerSymptomKeyList
    # len_devices_per_table[1]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 2 , 2)
    StateButtonList[_current_State+0] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+1] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+2] =  _TrimmerSymptomKeyList
    # len_devices_per_table[2]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 3 , 2)
    StateButtonList[_current_State+0] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+1] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+2] =  _TrimmerSymptomKeyList
    # len_devices_per_table[3]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 4 , 2)
    StateButtonList[_current_State+0] =  _CastingmachineSymptomKeyList
    StateButtonList[_current_State+1] =  _ElectricfurnaceSymptomKeyList    
    StateButtonList[_current_State+2] =  _ElectricfurnaceSymptomKeyList    
    StateButtonList[_current_State+3] =  _ElectricfurnaceSymptomKeyList    
    StateButtonList[_current_State+4] =  _ElectricfurnaceSymptomKeyList    
    # len_devices_per_table[4]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 5 , 2)
    StateButtonList[_current_State+0] =  _CastingmachineSymptomKeyList
    StateButtonList[_current_State+1] =  _ElectricfurnaceSymptomKeyList    
    StateButtonList[_current_State+2] =  _ElectricfurnaceSymptomKeyList    
    # len_devices_per_table[5]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 6 , 2)
    StateButtonList[_current_State+0] =  _CastingmachineSymptomKeyList
    StateButtonList[_current_State+1] =  _ElectricfurnaceSymptomKeyList    
    StateButtonList[_current_State+2] =  _ElectricfurnaceSymptomKeyList    
    # len_devices_per_table[6]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 7 , 2)
    StateButtonList[_current_State+0] =  _CuringwaterbathSymptomKeyList
    StateButtonList[_current_State+1] =  _CuringwaterbathSymptomKeyList    
    StateButtonList[_current_State+2] =  _SteamcleanerSymptomKeyList   
    StateButtonList[_current_State+3] =  _SteamcleanerSymptomKeyList   
    StateButtonList[_current_State+4] =  _CuringwaterbathSymptomKeyList
    # len_devices_per_table[7]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 8 , 2)
    StateButtonList[_current_State+0] =  _DispenserSymptomKeyList
    StateButtonList[_current_State+1] =  _DispenserSymptomKeyList    
    StateButtonList[_current_State+2] =  _SteamcleanerSymptomKeyList   
    StateButtonList[_current_State+3] =  _VacuummixerSymptomKeyList   
    StateButtonList[_current_State+4] =  _VacuummixerSymptomKeyList
    # len_devices_per_table[8]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 9 , 2)
    StateButtonList[_current_State+0] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+1] =  _VacuummixerSymptomKeyList    
    StateButtonList[_current_State+2] =  _DispenserSymptomKeyList
    StateButtonList[_current_State+3] =  _DispenserSymptomKeyList
    # len_devices_per_table[9]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 10 , 2)
    StateButtonList[_current_State+0] =  _TrimmerSymptomKeyList
    StateButtonList[_current_State+1] =  _VacuummixerSymptomKeyList    
    StateButtonList[_current_State+2] =  _SteamcleanerSymptomKeyList
    # len_devices_per_table[10]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 11 , 2)
    StateButtonList[_current_State+0] =  _PollcleanerSymptomKeyList
    StateButtonList[_current_State+1] =  _PollcleanerSymptomKeyList    
    StateButtonList[_current_State+2] =  _PollcleanerSymptomKeyList
    StateButtonList[_current_State+3] =  _PollcleanerSymptomKeyList
    # len_devices_per_table[11]?

    for i in range(len_tables)  :        
        pop_pushedStateList[nx_Child_in( nx_Child_in(first_4eng_State,1)+i ,3)] = True
    for i in range(len_tables)  :        
        push_StateList[nx_Child_in( nx_Child_in(first_4eng_State,1)+i ,1)] = True
        push_StateList[nx_Child_in( nx_Child_in(first_4eng_State,1)+i ,3)] = True
        for j in range(len(len_devices_per_table)) : 
            push_StateList[nx_Child_in( nx_Child_in(first_4eng_State,1)+i ,2)+j] = True


    fromStateMessageList[first_4eng_State] = SelectString+u'\n' 
    for i in range(len_tables)  :        
        fromStateMessageList[ nx_Child_in(first_4eng_State,1)+i ] = SelectString+u'\n'
        _current_State = nx_Child_in(first_4eng_State,1)+i
        fromStateMessageList[ nx_Child_in(_current_State,1) ] = SelectString+u'\n'  
        _current_State = nx_Child_in(_current_State,1)
        for j in range(len(len_devices_per_table)) : 
            fromStateMessageList[ nx_Child_in(_current_State,1)+j ] = SelectString+u'\n'
        _current_State = nx_Child_in(_current_State,1)
        fromStateMessageList[ nx_Child_in(_current_State,1) ] = SelectString+u'\n'            
        _current_State = nx_Child_in(_current_State,1)
        fromStateMessageList[ nx_Child_in(_current_State,1) ] = SelectString+SubmitString+u'\n'

    toStateMessageList[first_4eng_State] = AskTableNumberString
    for i in range(len_tables)  :        
        toStateMessageList[ nx_Child_in(first_4eng_State,1)+i ] = AskDeviceString
        _current_State = nx_Child_in(first_4eng_State,1)+i
        toStateMessageList[ nx_Child_in(_current_State,1) ] = DirectInsertDeviceString
        _current_State = nx_Child_in(_current_State,1)
        toStateMessageList[ nx_Child_in(_current_State,2) ] = DirectInsertSymptomString

    for i in range(len_tables)  :
        _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + i , 2)
        for j in range(len(len_devices_per_table)) :         
            toStateMessageList[_current_State+j] =  AskSymptomString


    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 0 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SandblastSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SandblastSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]  
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_VacuummixerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+3] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SandblastSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+4] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+5] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SandblastSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+6] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SandblastSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[0]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 1 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[1]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 2 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[2]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 3 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[3]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 4 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_CastingmachineSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+3] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+4] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    # len_devices_per_table[4]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 5 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_CastingmachineSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    # len_devices_per_table[5]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 6 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_CastingmachineSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_ElectricfurnaceSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    # len_devices_per_table[6]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 7 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_CuringwaterbathSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_CuringwaterbathSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SteamcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]   
    toSymptomStateMessageListsList[_current_State+3] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SteamcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]   
    toSymptomStateMessageListsList[_current_State+4] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_CuringwaterbathSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[7]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 8 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_DispenserSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_DispenserSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SteamcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]   
    toSymptomStateMessageListsList[_current_State+3] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_VacuummixerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]   
    toSymptomStateMessageListsList[_current_State+4] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_VacuummixerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[8]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 9 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_VacuummixerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_DispenserSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+3] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_DispenserSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[9]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 10 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_TrimmerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_VacuummixerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_SteamcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[10]?
    _current_State =  nx_Child_in( nx_Child_in(first_4eng_State,1) + 11 , 2)
    toSymptomStateMessageListsList[_current_State+0] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_PollcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+1] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_PollcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]    
    toSymptomStateMessageListsList[_current_State+2] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_PollcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    toSymptomStateMessageListsList[_current_State+3] =  [AskSymptomString, DirectInsertSymptomString, AskMultiSymptomString+u'\n'+_PollcleanerSymptomJoinString+u'\n\n'+ExplainSymptomInsertionString ]
    # len_devices_per_table[11]?

Error_NoInt     = 0x9
Error_NoSubTree = 0x8
def determineSubGraph( _State , _next=0 ) :
    if type( _State ) is not int :
        return  Error_NoInt
    else :
        num_str = format(_State,'02X')

        if len(num_str) < len( format(first_4work_State,'02X')) :
            return Error_NoSubTree
        else :
            if    num_str[:len(format(first_4work_State,'02X'))] ==  format(first_4work_State,'02X') :
                return nx_Child(first_4work_State,_next)
            elif  num_str[:len(format(first_3com_State,'02X'))] == format(first_3com_State,'02X') :
                return  nx_Child(first_3com_State, _next) 
            elif  num_str[:len(format(first_4eng_State,'02X'))] == format(first_4eng_State,'02X') :
                if _next == 0 :
                    return  first_4eng_State
                else :
                    return  nx_Child( first_4eng_State*0x10 +  int('0X'+num_str[len(format(first_4eng_State,'02X'))] ,0)  , _next -1 )                   
            elif  num_str[:len(format(first_3work_State,'02X'))] == format(first_3work_State ,'02X') :
                return  nx_Child(first_3work_State, _next)
            elif  num_str[:len(format(first_3handpiece_State,'02X'))] == format(first_3handpiece_State,'02X' ) :
                return  nx_Child(first_3handpiece_State, _next)
            else :
                return Error_NoSubTree


StateString = 'state'
LocationString = 'location'
SeatNumberString = 'seat number'
PartString = 'part'
SymptomString = 'symptom'
TimeString = 'time'

practice_sum_instance = { u'init' : [ ] }
sum_instance = { u'init' : [ ] } 

instance = { u'temp': {StateString:initial_State, 
                      LocationString    : u'',
                      SeatNumberString  : u'' ,
                      PartString        : u'',
                      SymptomString      : u'' 
                     }
}

IDString = 'ID'
NameString = 'Name'
GradeString = 'Grade'
RecordedYearString = 'RecordedYear'
InputModeString = 'InputMode'

organization = { u'init' :  { IDString : 0 ,
                              NameString : u'init' ,
                              GradeString :  1,
                              RecordedYearString : 2017,
                              InputModeString : 0 
                            }
}
temp_organization = { u'temp' :  { IDString : 0 ,
                                   NameString : u'temp' , 
                                   GradeString :  1,
                                   RecordedYearString : 2017,
                                   InputModeString : 0    
                                }
} 
org_rwfile_path = u'static/organization.txt'

def generateOrganization( _org)  :
    if os.path.exists(org_rwfile_path) :
        f = open( org_rwfile_path , 'r') 
        lines = f.readlines()
        for line in lines :
            tokens = line.decode('utf-8').split()
            if tokens[0] not in _org and \
                len(tokens) == 6 and \
                tokens[1].isdigit() and  tokens[3].isdigit() and  \
                tokens[4].isdigit() and  tokens[5].isdigit() :    # user_key  ID  Name   Grade    RecordedYear   InputMode
                _org[tokens[0]] = { IDString : int(tokens[1]) }
                _org[tokens[0]][NameString ] = tokens[2]
                _org[tokens[0]][GradeString ] = int(tokens[3])
                _org[tokens[0]][RecordedYearString ] = int(tokens[4])
                _org[tokens[0]][InputModeString ] = int(tokens[5])             
            elif tokens[0] not in _org and \
                len(tokens) == 4 and \
                tokens[1].isdigit() :    # user_key  ID  Name  InputMode
                _org[tokens[0]] = { IDString : int(tokens[1]) }
                _org[tokens[0]][NameString ] = tokens[2]
                _org[tokens[0]][GradeString ] = 5
                _org[tokens[0]][RecordedYearString ] = 2017
                _org[tokens[0]][InputModeString ] = int(tokens[3])             
            elif tokens[0] not in _org and \
                len(tokens) == 3 and \
                tokens[1].isdigit() :   # user_key  ID Name 
                _org[tokens[0]] = { IDString : int(tokens[1]) }
                _org[tokens[0]][NameString ] = tokens[2]
                _org[tokens[0]][GradeString ] = 1
                _org[tokens[0]][RecordedYearString ] = 2017
                _org[tokens[0]][InputModeString ] = 0 
        f.close()   

emailToOfficeList = []
emailForwardingList = []
emailAdminList = []
emailMulti_rofile_path = u'static/email2.txt'

CaptainList = {}
contact_rofile_path = u'static/contact_list.txt'
def generateMultiEmailToList( _email2list, _emailfowardinglist, _emailadminlist) :
    if os.path.exists(emailMulti_rofile_path) :
        f = open( emailMulti_rofile_path, 'r')
        lines = f.readlines()

        tokens = lines[0].split()
        for i in range(1,len(tokens)) : 
            _email2list.append(tokens[i])
        tokens = lines[1].split()
        for i in range(1,len(tokens)) : 
            _emailfowardinglist.append(tokens[i])
        tokens = lines[2].split()
        for i in range(1,len(tokens)) : 
            _emailadminlist.append(tokens[i])
        f.close()


def generateContactList( _CaptainList,  _OfficeEmailList, _ForwardingEmailList, _AdminEmailList) :
    if os.path.exists(contact_rofile_path) :
        f = open( contact_rofile_path, 'r')
        lines = f.readlines()
        tokens = lines[0].split()
        for i in range(1,len(tokens)) : 
            _OfficeEmailList.append(tokens[i])

        for j in range(1, 1+6) :
            tokens = lines[j].split()
            _CaptainList[j] = []
            for i in range(1,len(tokens)) : 
                #_CaptainList[j].append(tokens[i])
                #written in windows hangul , so it is euc-kr
                _CaptainList[j].append( unicode(  tokens[i] , 'euc-kr' ).encode('utf-8') )

        tokens = lines[7].split()
        for i in range(1,len(tokens)) : 
            _ForwardingEmailList.append(tokens[i])
        tokens = lines[8].split()
        for i in range(1,len(tokens)) : 
            _AdminEmailList.append(tokens[i])
        f.close()



gmailUserString = u'ID'
gmailPasswordString = u'Password'
gmailInformation = {}
emailFrom_rofile_path = u'static/document.txt'
def generateEmailFrom(_gmailInfo) :
    if os.path.exists(emailFrom_rofile_path) :
        f = open( emailFrom_rofile_path, 'r')
        f.readline()
        line = f.readline()
        tokens = line.split()
        if len(tokens) == 2 :
            _gmailInfo[gmailUserString] = tokens[0]
            _gmailInfo[gmailPasswordString] = tokens[1]
        else :
            _gmailInfo[gmailUserString] = "insufficient"
            _gmailInfo[gmailPasswordString] = "information"
        f.close()

pass_rofile_path = u'static/document.txt'

class  Arrow :
    def  __init__(self, mItemList = { } ) :
        self.mItemList = {   "message" : { "text": "" } ,    } 

    def  _make_Messages_change_State( self, _TextFlag, _text , _PhotoFlag  , _Photo , _mButtonFlag , _mButton,  _ButtonFlag , _fromState, _toState , _userRequest ) :
        self.mItemList["message"]["text"] = _text
        if _PhotoFlag   == True : 
            self.mItemList["message"]["photo"] = _Photo
        if _mButtonFlag == True : 
            self.mItemList["message"]["message_button"] = _mButton
        if _ButtonFlag  == True : 
            self.mItemList["keyboard"] = {  "type": "buttons" }
            self.mItemList["keyboard"]["buttons"] = StateButtonList[_toState]   
        instance[_userRequest['user_key']][StateString] = _toState

        if _fromState in push_StateList and \
           _fromState < _toState :
            if 'prev' in instance[_userRequest['user_key']] :
              instance[_userRequest['user_key']]['pprev'] = instance[_userRequest['user_key']]['prev'] 
            instance[_userRequest['user_key']]['prev']  = _fromState
        elif  _toState in pop_pushedStateList and \
              _fromState > _toState and \
            'pprev' in instance[_userRequest['user_key']] :
            instance[_userRequest['user_key']]['prev']  = instance[_userRequest['user_key']]['pprev']

        return jsonify(self.mItemList)

    def  _make_Message_Button_change_State( self, _TextFlag, _text , _ButtonFlag , _fromState ,_toState , _userRequest ) :
        return self._make_Messages_change_State(_TextFlag, _text , False , {} , False , {} , _ButtonFlag , _fromState ,_toState, _userRequest)

    def  make_Message_Button_change_State(self,  _fromState, _toState, _userRequest, _url_root=None)  :
        #_textMessage = _userRequest['content']+ fromStateMessageList[_fromState]  +   toStateMessageList[_toState]
        _textMessage = _userRequest['content']
        if _fromState in  fromStateMessageList :
            _textMessage += fromStateMessageList[_fromState]
        else :
            _textMessage += u'toState('+format(  _fromState , '#04x' )+u')' 

        if _toState in  toStateMessageList :
            _textMessage += toStateMessageList[_toState]
        else :
            _textMessage += u'toState('+format(  _toState , '#04x' )+u')' 

        _ButtonFlag = True
        _PhotoFlag  = True
        _Photo = {}
        if _toState in StateButtonList :
            _ButtonFlag = True 
        else :
            _ButtonFlag = False  
        if  _toState in StatePhotoList and \
            _url_root is not None :
            _PhotoFlag = True
            _Photo["url"] = _url_root + StatePhotoList[_toState]["url"] 
            _Photo["width"] = StatePhotoList[_toState]["width"]
            _Photo["height"] = StatePhotoList[_toState]["height"]
        else :
            _PhotoFlag = False     


        if _toState in StateButtonList and _toState in StateMultiChoiceList :
            if  _userRequest['user_key'] not in organization and _userRequest['user_key']  in temp_organization :
                _ButtonFlag = True
                _textMessage = _userRequest['content']+ fromStateMessageList[_fromState]  +   toSymptomStateMessageListsList[_toState][0]
            elif  _userRequest['user_key'] in organization and organization[ _userRequest['user_key'] ][InputModeString] == 1 :
                _ButtonFlag = False
                _textMessage = _userRequest['content']+ fromStateMessageList[_fromState]  +   toSymptomStateMessageListsList[_toState][1]
            elif  _userRequest['user_key'] in organization and organization[ _userRequest['user_key'] ][InputModeString] == 2 :
                _ButtonFlag = False
                _textMessage = _userRequest['content']+ fromStateMessageList[_fromState]  +   toSymptomStateMessageListsList[_toState][2]
            else   :
                _ButtonFlag = True
                _textMessage = _userRequest['content']+ fromStateMessageList[_fromState]  +   toSymptomStateMessageListsList[_toState][0]
                 


        return self._make_Messages_change_State(True, _textMessage , _PhotoFlag , _Photo  , False , {} , _ButtonFlag , _fromState, _toState, _userRequest)



def _isValidState(num) :
    if num in state :
        return True
    else:
        return False

def  _nx_Child(stage_num , score) :
    if score == 0 :
        return stage_num
    else :
        return _nx_Child(stage_num * 0x10 +1 , score-1)

def  nx_Child( stage_num , score ) :
    num = _nx_Child(stage_num, score)
    if _isValidState(num) :
        return num
    else :
        return  num*0x100 

def nx_Child_Sibling(stage_num , child_score, sibling_score) :
    num = nx_Child(stage_num, child_score) + sibling_score
    if _isValidState(num) :
        return num
    else :
        return  num*0x100 

def _prev_Parent(stage_num, score) :
    if score == 0 :
        return stage_num 
    else :
        return _prev_Parent( stage_num/0x10  , score-1)    

def prev_Parent(stage_num, score) :
    num = _prev_Parent(stage_num, score)
    if _isValidState(num) :
        return num
    else :
        return num * 0x100

def restore_prev_State(_UserRequestKey) :
    return instance[_UserRequestKey]['prev']

def isValidID( number ) :
    if number % 10000 in range(1, 80+1)  and \
       (number / 10000) % 100 == 74 :
       return True 
    else :
        return False
    return False 

class SummaryText :
    def  __init__(self, mText = '' , mTextGroup = {} , mTextGroup_separation = {} , mInstanceCount = 0 ) :
        self.mText = u'' 
        self.mTextGroup = {} 
        self.mTextGroup_separation = {}
        self.mInstanceCount = 0


    def  _generate(self, _TextMessage,_organization,_instance, _UserRequestKey, _key1=None, _OnlyPart=False) :
        self.mText += _TextMessage
        if _OnlyPart == False :
            self.mText += u'ID         :' + str(_organization[ _UserRequestKey ][IDString])+u'\n'
            self.mText += u'Name       :' + _organization[ _UserRequestKey ][NameString]+u'\n'
            if  GradeString in _organization[ _UserRequestKey ].keys() :
                self.mText += u'Grade       :' + str(_organization[ _UserRequestKey ][GradeString ])+u'\n'

        if _key1 is None :
            self.mText += u'location   :' + _instance[ _UserRequestKey ][LocationString]+u'\n'
            self.mText += u'seat number:' + _instance[ _UserRequestKey ][SeatNumberString]+u'\n'
            self.mText += u'part       :' + _instance[ _UserRequestKey ][PartString]+u'\n'
            self.mText += u'symptom    :' + _instance[ _UserRequestKey ][SymptomString]
        else : 
#            if _OnlyPart == False :
            self.mText += u'location   :' + _instance[ _UserRequestKey ][_key1][LocationString]+u'\n'
            self.mText += u'seat number:' + _instance[ _UserRequestKey ][_key1][SeatNumberString]+u'\n'
            self.mText += u'part       :' + _instance[ _UserRequestKey ][_key1][PartString]+u'\n'
            self.mText += u'symptom    :' + _instance[ _UserRequestKey ][_key1][SymptomString]+u'\n'
        return self.mText


    def  generateSumofAll(self, _organization,_instance, _UserRequestKey ) :

        self.mText += (datetime.now() + timedelta(hours=time_difference) )  .strftime("%Y-%m-%d %H:%M:%S") + u'\n'
        self.mText += u'최종 접수 예정:' +u'\n'

        if _UserRequestKey in _instance and  _UserRequestKey in _organization  :

            for i in range( len(_instance[_UserRequestKey]) ):
                if i == 0 :
                    self.mText += SummaryText()._generate(u'---------' + str(i+1) +  u'------------\n' , _organization, _instance, _UserRequestKey, i)
                else :
                    self.mText += SummaryText()._generate(u'---------' + str(i+1) +  u'------------\n' , _organization, _instance, _UserRequestKey, i, True)

            #subject = u'개인별고장 확인('+unicode ( (datetime.now() + timedelta(hours=time_difference)  ).strftime("%Y-%m-%d"))+u')'
            #mail(emailAdminList , subject , _textMessage.encode('utf-8'))

        if _UserRequestKey in practice_sum_instance and _UserRequestKey in _organization  :
            self.mText += u'\n'
            self.mText += u'(연습용)최종 접수 예정:' +u'\n'
            for i in range( len(practice_sum_instance[_UserRequestKey]) ):
                if i == 0 :
                    self.mText += SummaryText()._generate(u'---------' + str(i+1) +  u'------------\n' , _organization, practice_sum_instance, _UserRequestKey, i)
                else  :
                    self.mText += SummaryText()._generate(u'---------' + str(i+1) +  u'------------\n' , _organization, practice_sum_instance, _UserRequestKey, i, True)
        return self.mText


    def _genRegrouped2(self, _organization, _sum_instance) :

        self.mInstanceCount = 0
        _sumins_org_regrouping = {}
        for _userkey in _sum_instance.keys() :
            if _userkey not in _organization :
                continue
            for  _ins in  _sum_instance[_userkey]  : 
                if  _ins[LocationString]  not in  _sumins_org_regrouping :
                    _sumins_org_regrouping[ _ins[LocationString] ] = {}
                if  _ins[SeatNumberString] not in _sumins_org_regrouping[_ins[LocationString]] :
                    _sumins_org_regrouping[_ins[LocationString]][_ins[SeatNumberString]] = []
                if  _ins[LocationString]  not in self.mTextGroup :
                    self.mTextGroup[_ins[LocationString]] = u''
                _element = { 'user_key':_userkey }
                _element[IDString]     = _organization[_userkey][IDString]
                _element[NameString]   = _organization[_userkey][NameString]
                if GradeString in _organization[_userkey].keys() :
                    _element[GradeString ]   = _organization[_userkey][GradeString ]
                _element[SymptomString] = _ins[SymptomString]
                _element[PartString]   = _ins[PartString]
                _sumins_org_regrouping[_ins[LocationString]][_ins[SeatNumberString]].append(_element)                
        for _key0 in _sumins_org_regrouping :
#            self.mText += u'[['+ _key0    +u']]\n'
            for _key1 in _sumins_org_regrouping[_key0] :
                for _ins in _sumins_org_regrouping[_key0][_key1] :
                    self.mTextGroup[_key0] += _key1+u'번자리\t  '+_ins[PartString] + u'\t  '+_ins[SymptomString] +u'\n'
                    self.mInstanceCount += 1
        return self.mTextGroup

    def _genRegrouped3(self, _organization, _sum_instance) :

        self.mInstanceCount = 0
        _sumins_org_regrouping = {}
        for _userkey in _sum_instance.keys() :
            if _userkey not in _organization :
                continue
            for  _ins in  _sum_instance[_userkey]  : 
                if  _ins[LocationString]  not in  _sumins_org_regrouping :
                    _sumins_org_regrouping[ _ins[LocationString] ] = {}
                if  _ins[SeatNumberString] not in _sumins_org_regrouping[_ins[LocationString]] :
                    _sumins_org_regrouping[_ins[LocationString]][_ins[SeatNumberString]] = []
                if  _ins[LocationString]  not in self.mTextGroup :
                    self.mTextGroup[_ins[LocationString]] = []
                _element = { 'user_key':_userkey }
                _element[IDString]     = _organization[_userkey][IDString]
                _element[NameString]   = _organization[_userkey][NameString]
                if GradeString in _organization[_userkey].keys() :
                    _element[GradeString ]   = _organization[_userkey][GradeString ]
                _element[SymptomString] = _ins[SymptomString]
                _element[PartString]   = _ins[PartString]
                _sumins_org_regrouping[_ins[LocationString]][_ins[SeatNumberString]].append(_element)                
        for _key0 in _sumins_org_regrouping :
#            self.mText += u'[['+ _key0    +u']]\n'
            for _key1 in _sumins_org_regrouping[_key0] :
                for _ins in _sumins_org_regrouping[_key0][_key1] :
                    self.mTextGroup[_key0].append( _key1+u'번자리\t  '+_ins[PartString] + u'\t  '+_ins[SymptomString] +u'\n' )
                    self.mInstanceCount += 1
        return self.mTextGroup

    def _genRegrouped5(self, _organization, _sum_instance, _with=False) :

        self.mInstanceCount = 0
        _sumins_org_regrouping = {}
        for _userkey in _sum_instance.keys() :
            if _userkey not in _organization :
                continue
            
            for  _ins in  _sum_instance[_userkey]  : 
                if  _organization[_userkey][GradeString]     not in  _sumins_org_regrouping :
                    _sumins_org_regrouping[ _organization[_userkey][GradeString] ] = {}
                if  _ins[LocationString]  not in  _sumins_org_regrouping[ _organization[_userkey][GradeString] ] :
                    _sumins_org_regrouping[ _organization[_userkey][GradeString] ][ _ins[LocationString] ] = {}
                if  _ins[SeatNumberString] not in _sumins_org_regrouping[ _organization[_userkey][GradeString] ][ _ins[LocationString] ] :
                    _sumins_org_regrouping[ _organization[_userkey][GradeString] ][ _ins[LocationString] ][_ins[SeatNumberString]] = []

                if  _organization[_userkey][GradeString]  not in self.mTextGroup :
                    self.mTextGroup[ _organization[_userkey][GradeString] ] = {}
                if  _ins[LocationString] not in self.mTextGroup[ _organization[_userkey][GradeString] ] :
                    self.mTextGroup[ _organization[_userkey][GradeString] ][_ins[LocationString]] = u''

                if  _organization[_userkey][GradeString]  not in self.mTextGroup_separation :
                    self.mTextGroup_separation[ _organization[_userkey][GradeString] ] = {}
                if  _ins[LocationString] not in self.mTextGroup_separation[ _organization[_userkey][GradeString] ] :
                    self.mTextGroup_separation[ _organization[_userkey][GradeString] ][_ins[LocationString]] = []

                _element = { 'user_key':_userkey }
                _element[IDString]     = _organization[_userkey][IDString]
                _element[NameString]   = _organization[_userkey][NameString]
                if GradeString in _organization[_userkey].keys() :
                    _element[GradeString ]   = _organization[_userkey][GradeString ]
                _element[SymptomString] = _ins[SymptomString]
                _element[PartString]   = _ins[PartString]
                _sumins_org_regrouping[ _organization[_userkey][GradeString] ][_ins[LocationString]][_ins[SeatNumberString]].append(_element)                
        for _key0 in _sumins_org_regrouping :
            for _key1 in _sumins_org_regrouping[_key0] :
                for _key2 in _sumins_org_regrouping[_key0][_key1] :
                    for _ins in _sumins_org_regrouping[_key0][_key1][_key2] :
                        if  _key1 == _State1KeyList[3] :  # u'4.지하3층 실습실(핸드피스)' 
                            self.mTextGroup[_key0][_key1] += _key2+u'번장비\t  '
                        else :
                            self.mTextGroup[_key0][_key1] += _key2+u'번자리\t  '
                        self.mTextGroup[_key0][_key1] += _ins[PartString] + u'\t  '+_ins[SymptomString]
                        #self.mTextGroup[_key0][_key1] += _key2+u'번자리\t  '+_ins[PartString] + u'\t  '+_ins[SymptomString]
                        if _with == True  :
                            self.mTextGroup[_key0][_key1] += u'\t  '+  str(_ins[IDString]) + u' for System Maintenance'  
                        self.mTextGroup[_key0][_key1] += u'\n'

                        # need to upgrade and verify
                        if  _key1 == _State1KeyList[3] :  # u'4.지하3층 실습실(핸드피스)' 
                            self.mTextGroup_separation[_key0][_key1].append( _key2+u'번장비\t  '+_ins[PartString] + u'\t  '+_ins[SymptomString] +u'\n' )
                        else :
                            self.mTextGroup_separation[_key0][_key1].append( _key2+u'번자리\t  '+_ins[PartString] + u'\t  '+_ins[SymptomString] +u'\n' )

                        self.mInstanceCount += 1
        return self.mTextGroup

    def getTextGroup_Separation(self) :
        return self.mTextGroup_separation

    def getInstanceCount(self) :
        return self.mInstanceCount

    def showOrgFile(self)  :
        if os.path.exists(org_rwfile_path) :
            f = open( org_rwfile_path , 'r') 
            lines = f.readlines()
            for line in lines :
                self.mText += line.decode('utf-8')    
            f.close()      
        return self.mText

original_request_xlsx_file = u'static/original_request_office.xlsx'
target_request_xlsx_file          = u'static/target.xlsx'
target_request_xlsx_files =  [  u'static/target.xlsx', u'static/target1.xlsx' , u'static/target2.xlsx' , u'static/target3.xlsx' ,
                                u'static/target4.xlsx' , u'static/target5.xlsx' , u'static/target6.xlsx'   ] 

def mail( to, subject, body, attachlist=None):
    gmail_user = gmailInformation[gmailUserString]
    gmail_password = gmailInformation[gmailPasswordString]

    if  len(to) == 0 :
        to = emailAdminList

    if attachlist is None :    
        email_text = MIMEText( body , _charset="UTF-8")
    else :
        final_attachlist = []
        #attachlist = []
        for _element in attachlist :
            if  len(_element) == 2 :
                if  os.path.exists(_element[0])  :
                    fp = open(_element[0], 'rb')
                    part = MIMEBase('application', "octet-stream")
                    part.set_payload(fp.read())
                    fp.close()
                    encoders.encode_base64(part)
#                    part.add_header('Content-Disposition', 'attachment; filename="_target.xlsx"' )
                    part.add_header('Content-Disposition', 'attachment' , filename=_element[1] )
                    final_attachlist.append(part)


        email_text = MIMEMultipart()
        part1 = MIMEText( body , _charset="UTF-8")
        email_text.attach(part1)
        for _element in final_attachlist :
            email_text.attach(_element)


    email_text['Subject'] = subject
    email_text['From'] = gmail_user
    email_text['To'] = u", ".join(to)


    server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
    server.ehlo()
    server.login(gmail_user, gmail_password)
    server.sendmail(gmail_user , to, email_text.as_string() )

    server.close()


time_to_email = [ 8 ]  #8  ?
time_difference = 9
min2prepare = 10
def _calcTimer() :
    local_mail_hour = time_to_email[0]%24
    cloud_mail_hour = (24 + local_mail_hour -time_difference)%24
    x = datetime.today()
    y = x.replace(hour=cloud_mail_hour, minute=0, second=0, microsecond=0)
    if  datetime.now() +  timedelta(minutes=min2prepare) >= y :    
        y += timedelta(days=1)      
    #######    y=x.replace(day=x.day, hour=(x.hour+1)%24, minute=0, second=0, microsecond=0)    ### for test
    delta_t=y-x
    secs=delta_t.seconds+1
    return secs

requester_name = u'강신청'
requester_phone = u'010-0001-0005'

_GradeNumToStr = {  1:u'예1',
                    2:u'예2',
                    3:u'본1',
                    4:u'본2',
                    5:u'본3',
                    6:u'본4'
}

def GradeNumberToString(num) :
    if type(num) != int :
        return u'Unknown'
    elif num not in _GradeNumToStr.keys() :
        return u'Unknown'
    else :
        return _GradeNumToStr[num]        

class MailBodyandAttachment :
    def  __init__(self, mBody = u'' , mAttachmentList = [] , mInstanceCount = 0, mOfficeMailingList=[]) :
        self.mBody = u'' 
        self.mAttachmentList = []  
        self.mInstanceCount = 0
        self.mOfficeMailingList = []
    def prepare(self) :

        s = SummaryText()
        _TextGroup = s._genRegrouped2( organization, sum_instance )
        if  len(_TextGroup.keys()) == 0 :
            return False
        self.mInstanceCount = s.getInstanceCount()

        shutil.copy( original_request_xlsx_file,target_request_xlsx_file)
        excel_document = load_workbook(filename = target_request_xlsx_file)
        source = excel_document.get_sheet_by_name('Sheet1')
        #source = excel_document.active
        source['B5'] = unicode ( (datetime.now() + timedelta(hours=time_difference)  ).strftime("%Y-%m-%d")  )
        source['D5'] = requester_name 
        source['F5'] = requester_phone             

        for _key in  _TextGroup.keys() :
            self.mBody += u'[['+ _key  +u']]\n'
            self.mBody += _TextGroup[_key]          

            target = excel_document.copy_worksheet( source )
            target.title = _key[2:]  
            target['B4'] = _key[2:] 
            target['A7'] = _TextGroup[_key]
        excel_document.remove_sheet(source)
        excel_document.save(filename = target_request_xlsx_file)

        daily_filename = u'request('+unicode ( (datetime.now() + timedelta(hours=time_difference)  ).strftime("%Y-%m-%d")  )+u').xlsx'.encode('utf-8')
        self.mAttachmentList.append( [target_request_xlsx_file , daily_filename ] )
        self.mAttachmentList.append( [ u'static/images/4work_seats.png', u'B403_Seats.png'.encode('utf-8')] )
        self.mAttachmentList.append( [ u'static/images/4work_seats.png', u'B303_Seats.png'.encode('utf-8')] )
        self.mAttachmentList.append( [ u'static/images/3com_seats.png',  u'B306_Seats.png'.encode('utf-8')] )
        self.mAttachmentList.append( [ u'static/images/4eng_tables.png', u'B404_Tables.png'.encode('utf-8')] )
        return True

    def prepare5(self, _with=False) :

        s = SummaryText()
        if _with == False :
            _TextGroup = s._genRegrouped5( organization, sum_instance )
        else :
            _TextGroup = s._genRegrouped5( organization, sum_instance , True )          
        self.mOfficeMailingList += emailToOfficeList  
        if  len(_TextGroup.keys()) == 0 :
            return False
        self.mInstanceCount = s.getInstanceCount()

        for _key0 in  _TextGroup.keys() :
            shutil.copy( original_request_xlsx_file, target_request_xlsx_files[_key0])
            excel_document = load_workbook(filename = target_request_xlsx_files[_key0])
            source = excel_document.get_sheet_by_name('Sheet1')
            source['B5'] = unicode ( (datetime.now() + timedelta(hours=time_difference)  ).strftime("%Y-%m-%d")  )
            source['D5'] = CaptainList[_key0][1]    #captain name 
            source['F5'] = CaptainList[_key0][2]    #captain phone number
            self.mBody += u'[['+ GradeNumberToString(_key0)  +u']]'+u'\n'
                                                      
            for _key1 in  _TextGroup[_key0].keys() :
                self.mBody += u'['+ _key1 +u']'+u'\n'
                self.mBody += _TextGroup[_key0][_key1]          
                              
                target = excel_document.copy_worksheet( source )
                target.title = _key1[2:]  + u'('+ GradeNumberToString(_key0) +u')'
                target['B4'] = _key1[2:]
                target['A7'] = _TextGroup[_key0][_key1]

            excel_document.remove_sheet(source)
            excel_document.save(filename = target_request_xlsx_files[_key0])

            daily_filename = u'request(' + unicode ( (datetime.now() + timedelta(hours=time_difference)  ).strftime("%Y-%m-%d") )+u')(Grade'+ str(_key0) +u').xlsx'.encode('utf-8')
            self.mAttachmentList.append( [target_request_xlsx_files[_key0] , daily_filename ] )
            self.mOfficeMailingList.append( CaptainList[_key0][0] )
        self.mAttachmentList.append( [ u'static/images/4work_seats.png', u'B403_Seats.png'.encode('utf-8')] )
        self.mAttachmentList.append( [ u'static/images/4work_seats.png', u'B303_Seats.png'.encode('utf-8')] )
        self.mAttachmentList.append( [ u'static/images/3com_seats.png',  u'B306_Seats.png'.encode('utf-8')] )
        self.mAttachmentList.append( [ u'static/images/4eng_tables.png', u'B404_Tables.png'.encode('utf-8')] )
        
        return True


    def getBody(self) :
        return self.mBody
    def getAttachmentList(self) :
        return self.mAttachmentList
    def getInstanceCount(self) :
        return self.mInstanceCount  
    def getOfficeMailingList(self) :
        return self.mOfficeMailingList

def Org2File( _organization, _file) :
    f = open( _file , 'w')
    for key in _organization.keys() :
        _line = key  + u'  '+ str(_organization[key][IDString]) + u'  '+ _organization[key][NameString] 
        if GradeString in _organization[key].keys() : 
            _line += u'  '+ str(_organization[key][GradeString]) 
        if RecordedYearString in _organization[key].keys() : 
            _line += u'  '+ str(_organization[key][RecordedYearString]) 
        if InputModeString in _organization[key].keys() : 
            _line += u'  '+ str(_organization[key][InputModeString]) 
        _line += u'\n'
        if  isinstance(_line, unicode) :
            f.write(_line.encode('utf-8')) 
    f.close()    

def periodic_mail_forwarding()  :   
    try :
        m = MailBodyandAttachment()
        m.prepare5()

    except :
        mail( emailAdminList, u'periodic_mail_case1', u'check'.encode('utf-8'))
    try :
        #this is yesterday's summary
        #subject = u'실습실 수리 신청입니다('+unicode((datetime.now()+timedelta(hours=time_difference)-timedelta(days=1)).strftime("%Y-%m-%d"))+u')' 
        subject = u'실습실 수리 신청입니다(총 '+str(m.getInstanceCount())+u'건)('
        subject += unicode( (datetime.now()+timedelta(hours=time_difference) ).strftime("%Y-%m-%d") )+u')' 
        mail(  m.getOfficeMailingList(), subject , m.getBody().encode('utf-8') , m.getAttachmentList() )

    except Exception as ex :
#        mail( emailAdminList, u'periodic_mail_case2', u'check'.encode('utf-8'))
        mail( emailAdminList, u'periodic_mail_case2', str(ex).encode('utf-8'))

    try :
        n = MailBodyandAttachment()
        n.prepare5(True)
        subject = u'Maintenance실습실 수리 신청입니다(총 '+str(m.getInstanceCount())+u'건)('
        subject += unicode( (datetime.now()+timedelta(hours=time_difference) ).strftime("%Y-%m-%d") )+u')' 
        mail(  emailAdminList, subject , m.getBody().encode('utf-8') , m.getAttachmentList() )

    except Exception as ex :
#        mail( emailAdminList, u'periodic_mail_case2', u'check'.encode('utf-8'))
        mail( emailAdminList, u'periodic_mail_case2.5', str(ex).encode('utf-8'))

    try :
        sum_instance.clear()  #clear sum_instance 
        Org2File(organization, org_rwfile_path)  # organization to organization file
        practice_sum_instance.clear()

        _textMessage = u'sum of instances'+ u'\n'
        _textMessage += str(sum_instance) + u'\n'
        _textMessage += SummaryText().showOrgFile()
        mail(emailAdminList, u'after periodic mail', _textMessage.encode('utf-8') )  #check sum_instance and organization file
    except :
        mail( emailAdminList, u'periodic_mail_case3', u'check'.encode('utf-8'))
    try :
        t = Timer( _calcTimer(), periodic_mail_forwarding)
        t.start()
    except :
        mail( emailAdminList, u'periodic_mail_case4', u'check'.encode('utf-8'))


def hello_world() :
    try : 
        s = Timer( _calcTimer() , periodic_mail_forwarding)
        s.start()
    except :
        mail( emailAdminList, u'hello_world_case0', u'check'.encode('utf-8'))


generate4EngStatesInformation()
generateOrganization(organization)
#generateMultiEmailToList(emailToOfficeList, emailForwardingList, emailAdminList)
generateContactList( CaptainList, emailToOfficeList, emailForwardingList, emailAdminList)
generateEmailFrom(gmailInformation)
hello_world()

real_number_list = {}    ## numbers(index) to delete.  numbers will be sorted by ascending order 

PrevTimeString = u'prev'
MessageTime = {}

@app.route('/')
def Welcome():
    return app.send_static_file('index.html')

@app.route('/myapp')
def WelcomeToMyapp():
    return 'Welcome again to my app running on Bluemix!!'

@app.route('/keyboard')
def Keyboard():
    try :    
        ItemList = {
            'type': 'buttons', 'buttons' : StateButtonList[initial_State]
        }    
        return jsonify(ItemList)
    except Exception as ex :
        log = str(ex)
        ItemList = {
            'type': 'buttons', 'buttons' : [ log ]
        }    
        return jsonify(ItemList)


@app.route('/message', methods=['POST'])
def GetMessage():
    try :
        userRequest = json.loads(request.get_data()) 

        # if its a 1st message, you have to make instance
        if userRequest['user_key'] not in instance :
            instance[userRequest['user_key']] = { StateString :  initial_State }
        # if its a sudden quit-and-reenter case, then make state initial
        if  instance[userRequest['user_key']][StateString] != initial_State and \
            userRequest['content']  in  StateButtonList[initial_State] :
                instance[userRequest['user_key']][StateString] = initial_State        

        currentState = instance[userRequest['user_key']][StateString]
        currentTime = datetime.now()

        if userRequest['user_key'] not in MessageTime :
            MessageTime[userRequest['user_key']] = { }
            if  PrevTimeString not in MessageTime  :
                MessageTime[userRequest['user_key']][PrevTimeString] = currentTime
        else :
            if  PrevTimeString not in MessageTime[userRequest['user_key']]  :
                MessageTime[userRequest['user_key']][PrevTimeString] = currentTime
            else  :
                if  MessageTime[userRequest['user_key']][PrevTimeString] + timedelta(milliseconds=800) > currentTime :
                    MessageTime[userRequest['user_key']][PrevTimeString] = currentTime
                    _textMessage =  SameButtonString+hex(currentState)
                    return Arrow().make_Message_Button_change_State(currentState, currentState, userRequest, request.url_root)            
    #                return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,  currentState, currentState, userRequest)          
                else  :
                    MessageTime[userRequest['user_key']][PrevTimeString] = currentTime


        #select initially 
        if instance[userRequest['user_key']][StateString] ==  initial_State  :        #state 1
            currentState = instance[userRequest['user_key']][StateString]   
            if  userRequest['content']  ==  StateButtonList[ currentState ][0] :
                if  userRequest['user_key'] not in organization :
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,1) ,userRequest)
                else :
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,4) ,userRequest)
                    ##return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,3) ,userRequest)
            elif userRequest['content']  ==  StateButtonList[ currentState ][1] :
                _textMessage = userRequest['content']+SelectString+u'\n'+u'KnP Version: '+ VersionString +u'\n'
                _textMessage += SummaryText().generateSumofAll(organization, sum_instance, userRequest['user_key'] ) 
                #return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,  currentState, currentState, userRequest)
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,  currentState, nx_Child_Sibling(currentState,1,1)   , userRequest)
            elif userRequest['content']  ==  StateButtonList[ currentState ][2] :             
                _textMessage = userRequest['content']+SelectString+u'\n\n'
                _textMessage += u'KnP 고장수리 톡 사용안내 1부\n(처음 사용하기)'+u'\n'
                _textMessage += u'https://youtu.be/OCy3ylSq-hs'+u'\n\n'
                _textMessage += u'KnP 고장수리 톡 사용안내 2부\n(입력 예시)'+u'\n'
                _textMessage += u'https://youtu.be/rGUweN0JcBg'+u'\n\n'
                _textMessage += u'KnP 고장수리 톡 사용안내 3부\n(삭제하기)'+u'\n'
                _textMessage += u'https://youtu.be/NJOEiSimICY'
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,currentState , currentState, userRequest)
            elif userRequest['content']  ==  StateButtonList[ currentState ][3] :            
                return Arrow().make_Message_Button_change_State(currentState,    nx_Child_Sibling(currentState,1,3) , userRequest)
                  
            else :
                _textMessage = userRequest['content']+SelectString+u'\n'+UnderConstructionString
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,currentState , currentState, userRequest)

        elif   instance[userRequest['user_key']][StateString] \
            in [ nx_Child(initial_State,1) , first_Independent_IDInsert_State ]  :    #11, 141
            currentState = instance[userRequest['user_key']][StateString]   
            try :
                if isValidID( int ( userRequest['content'] ) ) :
                    temp_organization[userRequest['user_key']] = { IDString : int ( userRequest['content'] ) }                
                    return Arrow().make_Message_Button_change_State( currentState, nx_Child( currentState, 1 )  , userRequest)
                elif  int ( userRequest['content'] ) == 0 : # return to prev menu
                    return Arrow().make_Message_Button_change_State( currentState ,  prev_Parent(currentState,1)  , userRequest)
                else : 
                    _textMessage = userRequest['content']+SelectString+u'\n' + InsertValidNumberString
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, False, currentState, currentState, userRequest)             
            except ValueError :
                _textMessage = userRequest['content']+SelectString+u'\n'+ InsertNumberString
                return Arrow()._make_Message_Button_change_State(True, _textMessage, False,currentState , currentState , userRequest)    

        #insert Name and prepare arrows
        elif   instance[userRequest['user_key']][StateString]  \
            in [  nx_Child(initial_State,2),   nx_Child(first_Independent_IDInsert_State,1) ]  :               #111 , 1411
            currentState = instance[userRequest['user_key']][StateString] 
            if  userRequest['content']  == '0' :
                return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest, request.url_root)   
            else :
                temp_organization[userRequest['user_key']][NameString] = userRequest['content'] 
                temp_organization[userRequest['user_key']][InputModeString ] = 0     # intput mode is set as default (0) 
                return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,1) , userRequest ) 

        # insert Grade
        elif   instance[userRequest['user_key']][StateString]  \
            in [  nx_Child(initial_State,3),   nx_Child(first_Independent_IDInsert_State,2) ]  :               #111 , 1411
            currentState = instance[userRequest['user_key']][StateString] 
            if     userRequest['content']  == StateButtonList[ currentState ][6] :
                return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest, request.url_root)               
            elif   userRequest['content']  == StateButtonList[ currentState ][5]  :
                temp_organization[userRequest['user_key']][GradeString ] = 1
            elif   userRequest['content']  == StateButtonList[ currentState ][4]  :
                temp_organization[userRequest['user_key']][GradeString ] = 2
            elif   userRequest['content']  == StateButtonList[ currentState ][3]  :
                temp_organization[userRequest['user_key']][GradeString ] = 3
            elif   userRequest['content']  == StateButtonList[ currentState ][2]  :
                temp_organization[userRequest['user_key']][GradeString ] = 4
            elif   userRequest['content']  == StateButtonList[ currentState ][1]  :
                temp_organization[userRequest['user_key']][GradeString ] = 5
            elif   userRequest['content']  == StateButtonList[ currentState ][0]  :
                temp_organization[userRequest['user_key']][GradeString ] = 6
            else :
                temp_organization[userRequest['user_key']][GradeString ] = 1
            temp_organization[userRequest['user_key']][RecordedYearString ] = datetime.now().year

            if currentState == nx_Child(first_Independent_IDInsert_State,2) :
                _textMessage = userRequest['content']+SelectString+u'\n'+  LastYesNoString +u'\n'
                _textMessage += u'ID   :'+ str(temp_organization[userRequest['user_key']][IDString])+u'\n'
                _textMessage += u'Name :'+ temp_organization[userRequest['user_key']][NameString]+u'\n'
                _textMessage += u'Grade :'+ str(temp_organization[userRequest['user_key']][GradeString ] )
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  nx_Child( currentState ,1) , userRequest)             
            else :
                return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,1) , userRequest ) 


        #process decision for Go-to-Initial, Delete some, Delte All, Goto  Prev
        elif   instance[userRequest['user_key']][StateString]  \
            in [  nx_Child_Sibling(initial_State,1,1) ]  :               #12
            currentState = instance[userRequest['user_key']][StateString] 
            _UserRequestKey = userRequest['user_key']
            if     userRequest['content']  in [   StateButtonList[ currentState ][1] ,  StateButtonList[ currentState ][2] ] :  #  Delete some or all
                if   _UserRequestKey not in organization  or \
                     _UserRequestKey not in sum_instance.keys()  or  \
                     len(sum_instance[_UserRequestKey]) == 0 :
                    _textMessage = userRequest['content']+SelectString+u'\n'
                    _textMessage += u'삭제할 item이 없습니다'+u'\n'
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  prev_Parent(currentState,1) , userRequest)
                elif   userRequest['content'] ==  StateButtonList[ currentState ][1]  :     #delete some
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,1) , userRequest)
                elif   userRequest['content'] ==  StateButtonList[ currentState ][2]  :     #delete all
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child_Sibling(currentState,1,1) , userRequest)
                else  :  #which case , defensive else                                    
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,1) , userRequest)

            elif  userRequest['content']  in  [    StateButtonList[currentState][0] ,  StateButtonList[currentState][3] ] :
                return Arrow().make_Message_Button_change_State(currentState,  prev_Parent(currentState,1) , userRequest)

        #process item numbers to delete
        elif   instance[userRequest['user_key']][StateString]  \
            in [  nx_Child(nx_Child_Sibling(initial_State,1,1),1) ]  :               #121
            currentState = instance[userRequest['user_key']][StateString]
            _UserRequestKey = userRequest['user_key'] 

            number_list = [] 
            real_number_list[_UserRequestKey] = []  #initialize real_number_list per userkey
            if  _UserRequestKey in sum_instance.keys()  and  len(sum_instance[_UserRequestKey]) != 0 :
                preserve_real_number_list= list ( range( len(sum_instance[_UserRequestKey]) )   )  
            tokens = re.split(r'(\s*\,\s*|\s+)', userRequest['content'] )
            if  len(tokens) == 1 and \
                tokens[0].strip().isdigit() and \
                int ( tokens[0].strip() ) == 0 :

                    _textMessage = userRequest['content']+SelectString+u'\n'+u'KnP Version: '+ VersionString +u'\n'
                    _textMessage += SummaryText().generateSumofAll(organization, sum_instance, userRequest['user_key'] ) 
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,  currentState, prev_Parent(currentState,1) , userRequest)                
                    #return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest)              
            for token  in  tokens  :
                if token.strip().isdigit() :   # if token is digit
                    if int ( token.strip() ) not in number_list :   # and token is not in number_list
                        number_list.append(  int ( token.strip() )  )    #append token
                    else :                                          # token is already in number_list
                        continue                                    # skip it
                elif bool (   re.match( r'\s*\,\s*|\s+' , token ) ) :   # if token is   whitespace or ,
                    continue                                            # skip it
                else  :                                                 # if token is char 
                    _textMessage = token+SelectString+u'\n'+ InsertNumberString               
                    return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState, currentState , userRequest)   # notify it and return back
            if  len ( number_list ) > 1 :
                number_list.sort()  #sort numbers

            if  len ( number_list ) > 0 and number_list[0] == 0  :
                _textMessage = str(number_list[0]) +SelectString+u'\n'+ InsertValidNumberString               
                return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState, currentState , userRequest)   # notify it and return back
            if  len ( number_list ) > 0 and  _UserRequestKey  in organization and  _UserRequestKey in sum_instance and number_list[-1] >  len(sum_instance[_UserRequestKey]) :
                _textMessage = str(number_list[-1])+SelectString+u'\n'+ InsertValidNumberString               
                return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState, currentState , userRequest)   # notify it and return back

            for number in number_list :
                real_number_list[_UserRequestKey].append( number-1 )
            for number in real_number_list[_UserRequestKey] :
                if  number in preserve_real_number_list  :
                    preserve_real_number_list.remove(number)
            _textMessage = userRequest['content']+SelectString+u'\n'
            _textMessage += AskDeletionString +u'\n'
            _textMessage += u'\n'
            _textMessage += u'삭제할 item들:' +u'\n'
            for i in real_number_list[_UserRequestKey] :  # for this print, real_number_list must be orderd by ascending order
                _textMessage += SummaryText()._generate(u'---------' + str(i+1) +  u'------------\n' , organization, sum_instance, _UserRequestKey, i)
            _textMessage += u'\n'
            _textMessage += u'남길 item들:' +u'\n'
            for i in preserve_real_number_list  :
                _textMessage += SummaryText()._generate(u'---------' + str(i+1) +  u'------------\n' , organization, sum_instance, _UserRequestKey, i)
            return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,  currentState, nx_Child(currentState,1) , userRequest)                

        # decide to delete (Y/N/Prev)
        elif   instance[userRequest['user_key']][StateString]  \
            in [  nx_Child(nx_Child_Sibling(initial_State,1,1),2) ]  :               #1211
            currentState = instance[userRequest['user_key']][StateString]   
            _UserRequestKey = userRequest['user_key'] 
            if  userRequest['content']  ==  StateButtonList[ currentState ][0] :
                _revorder_real_number_list = sorted( real_number_list[_UserRequestKey], reverse=True  )
                for number in  _revorder_real_number_list   :
                    if  _UserRequestKey in sum_instance.keys()  and  number in range( len(sum_instance[_UserRequestKey])  )  : 
                        del sum_instance[_UserRequestKey][number] 
                if   _UserRequestKey in sum_instance.keys() and len(sum_instance[_UserRequestKey])   == 0  :
                    sum_instance.pop( _UserRequestKey )

                _textMessage = userRequest['content']+SelectString+u'\n'+u'KnP Version: '+ VersionString +u'\n'
                _textMessage += SummaryText().generateSumofAll(organization, sum_instance, userRequest['user_key'] ) 
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,  currentState, prev_Parent(currentState,2) , userRequest)          

            elif userRequest['content']  ==  StateButtonList[ currentState ][2] :
                return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest)

            else : 
                return  Arrow().make_Message_Button_change_State(currentState,initial_State, userRequest)

        # decide to delete All(Y/N/Prev)
        elif   instance[userRequest['user_key']][StateString]  \
            in [  nx_Child_Sibling(nx_Child_Sibling(initial_State,1,1),1,1) ]  :               #122
            currentState = instance[userRequest['user_key']][StateString]   
            _UserRequestKey = userRequest['user_key'] 
            if  userRequest['content'] == StateButtonList[ currentState ][0]   :
                if   _UserRequestKey  in organization  and  _UserRequestKey in sum_instance.keys()   :
                    sum_instance.pop( _UserRequestKey )
                return  Arrow().make_Message_Button_change_State(currentState,initial_State, userRequest)
            elif  userRequest['content'] == StateButtonList[ currentState ][1]   :
                return  Arrow().make_Message_Button_change_State(currentState,initial_State, userRequest)
            else :
                _textMessage = userRequest['content']+SelectString+u'\n'+u'KnP Version: '+ VersionString +u'\n'
                _textMessage += SummaryText().generateSumofAll(organization, sum_instance, userRequest['user_key'] ) 
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True ,  currentState, prev_Parent(currentState,1) , userRequest)                

        elif   instance[userRequest['user_key']][StateString]  == nx_Child( first_Independent_IDInsert_State,3) : #14111
            currentState = instance[userRequest['user_key']][StateString]   
            if  userRequest['content']  ==  StateButtonList[ currentState ][0] :
                if  userRequest['user_key'] in temp_organization :

                    #if it is first ID/Nmae/Grade Inserttion, make own organization room( organization[userRequest['user_key']] )
                    if  userRequest['user_key'] not in  organization :
                        organization[userRequest['user_key']] = {}    
                    #update ID/Name/GradeString, whether org[ID/Name/GradeString] is already saved or not. that is becasue this can be whether first insertion or modification    
                    organization[userRequest['user_key']][IDString] = temp_organization[userRequest['user_key']][IDString]     
                    organization[userRequest['user_key']][NameString] = temp_organization[userRequest['user_key']][NameString]   
                    if  GradeString in temp_organization[userRequest['user_key']]  :
                        organization[userRequest['user_key']][GradeString ] = temp_organization[userRequest['user_key']][GradeString ]   

                    if  RecordedYearString in temp_organization[userRequest['user_key']] and RecordedYearString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][RecordedYearString ] = temp_organization[userRequest['user_key']][RecordedYearString ]   
                    if  InputModeString in temp_organization[userRequest['user_key']] and  InputModeString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][InputModeString] = temp_organization[userRequest['user_key']][InputModeString]   

                    temp_organization.pop( userRequest['user_key'] ,  None)
                    
                _textMessage = userRequest['content']+  u'\n' +SubmitString 
                instance[userRequest['user_key']] = { StateString : initial_State }            
                return  Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState, initial_State, userRequest)
            elif userRequest['content']  ==  StateButtonList[ currentState ][2] :
                return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest)
            else : 
                _textMessage = userRequest['content']+ u'\n'+ CancelString 
                instance[userRequest['user_key']] = { StateString : initial_State }            
                return  Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,initial_State, userRequest)

        #select location  
        ##elif instance[userRequest['user_key']][StateString] == nx_Child( initial_State,3) :  #1111
        elif instance[userRequest['user_key']][StateString] == nx_Child( initial_State,4) :  #1111
            currentState = instance[userRequest['user_key']][StateString]
            if  userRequest['content']  ==  StateButtonList[currentState][0] :
                instance[userRequest['user_key']][LocationString] = userRequest['content']
                return Arrow().make_Message_Button_change_State(currentState, first_4work_State, userRequest, request.url_root  )
            elif userRequest['content']  ==  StateButtonList[currentState][1] :
                instance[userRequest['user_key']][LocationString] = userRequest['content']
                return Arrow().make_Message_Button_change_State(currentState, first_4eng_State, userRequest, request.url_root  )
            elif userRequest['content']  ==  StateButtonList[currentState][2] :            
                instance[userRequest['user_key']][LocationString] = userRequest['content']
                return Arrow().make_Message_Button_change_State(currentState, first_3work_State  , userRequest , request.url_root )
            elif userRequest['content']  ==  StateButtonList[currentState][3] :
                instance[userRequest['user_key']][LocationString] = userRequest['content']
                return Arrow().make_Message_Button_change_State(currentState, first_3handpiece_State, userRequest, request.url_root  )
            elif userRequest['content']  ==  StateButtonList[currentState][4] :
                instance[userRequest['user_key']][LocationString] = userRequest['content']
                return Arrow().make_Message_Button_change_State(currentState, first_3com_State, userRequest, request.url_root )
            elif userRequest['content']  ==  StateButtonList[currentState][5] :  # return to prev menu
                if userRequest['user_key'] in organization :
                  return Arrow().make_Message_Button_change_State( currentState , prev_Parent(currentState,4) , userRequest)
                  ##return Arrow().make_Message_Button_change_State( currentState , prev_Parent(currentState,3) , userRequest)
                else :
                  ##return Arrow().make_Message_Button_change_State( currentState,  prev_Parent(currentState,2) , userRequest  )            
                  return Arrow().make_Message_Button_change_State( currentState,  prev_Parent(currentState,1) , userRequest  )            
            else :
                _textMessage = userRequest['content']+SelectString+u'\n'+'(state:'+ str(instance[userRequest['user_key']][StateString]) + ')'
                instance[userRequest['user_key']] = { StateString : initial_State }            
                return  Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState ,  initial_State, userRequest)

        #select memebership item and prepare arrows
        elif instance[userRequest['user_key']][StateString] == nx_Child_Sibling( initial_State ,1,3) :
            currentState = instance[userRequest['user_key']][StateString]
            if  userRequest['content']  ==  StateButtonList[currentState][0] :
                if  userRequest['user_key'] in organization :
                    key = userRequest['user_key']
                    _textMessage  = u'이미 입력된 ID와 Name이 있습니다.'+u'\n'
                    _textMessage += u'기존 ID   :'+ str(organization[key][IDString])+u'\n'
                    _textMessage += u'기존 Name :'+ organization[key][NameString]+u'\n'
                    if  GradeString in  organization[userRequest['user_key']]  :
                        _textMessage += u'Grade :'+ str(organization[key][GradeString ])+u'\n'
                    _textMessage += u'\n'
                    _textMessage += toStateMessageList[first_Independent_IDInsert_State]+u'\n'
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, False, currentState, first_Independent_IDInsert_State , userRequest )            

                    #fill this later. for one user multi device case
                    #elif organization[key]['ID'] == temp_organization[userRequest['user_key']]['ID'] :
                        #_ItemList["message"]["text"] =  u'다른 device를 사용 중입니다.' +  u'\n'
                        #_ItemList["message"]["text"] += u'기존 ID   :'+ str(organization[key]['ID'])+u'\n'
                        #_ItemList["message"]["text"] += u'기존 Name :'+ organization[key]['Name']+u'\n'
                        #_ItemList["message"]["text"] += u'새로운 device를 사용하시겠습니까?.'
                        #_ItemList["keyboard"]["buttons"] =    StateButtonList[41111] 
                        #instance[userRequest['user_key']]['state'] = state[41111] 
                        #return jsonify(_ItemList)
                else :
                    return Arrow().make_Message_Button_change_State(currentState, first_Independent_IDInsert_State , userRequest)

            elif  userRequest['content']  ==  StateButtonList[currentState][1] :
                if userRequest['user_key'] in organization:
                    _textMessage = userRequest['content']+SelectString + u'\n' + AskDeletionString+ u'\n'
                    _textMessage += u'ID   :'+ str(organization[userRequest['user_key']][IDString])+u'\n'
                    _textMessage += u'Name :'+ organization[userRequest['user_key']][NameString]+u'\n'
                    if  GradeString in  organization[userRequest['user_key']]  :
                        _textMessage += u'Grade :'+ str(organization[userRequest['user_key']][GradeString ])+u'\n'
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState, nx_Child_Sibling(currentState,1,1) , userRequest )            
                else : 
                    _text_ =  userRequest['content']+SelectString + u'\n' + u'등록된 ID와 Name이 없습니다.'           
                    return Arrow()._make_Message_Button_change_State( True, _text_, True, currentState,  initial_State , userRequest ) 
            elif  userRequest['content']  ==  StateButtonList[currentState][2] :   # go to Input Mode Selection
                currentState = instance[userRequest['user_key']][StateString]
                if userRequest['user_key'] not in organization :
                    _textMessage =  userRequest['content']+SelectString + u'\n' + u'등록된 ID와 Name이 없습니다.'           
                    return Arrow()._make_Message_Button_change_State( True, _textMessage, True, currentState,  initial_State , userRequest ) 
                else  :
#                    _textMessage =  userRequest['content']+SelectString + u'\n'
                    _textMessage  =  userRequest['content'] + fromStateMessageList[currentState]
                    _textMessage +=  toStateMessageList[nx_Child_Sibling(currentState,1,2) ] + u'\n\n'
                    if  InputModeString in  organization[userRequest['user_key']]  :
                        _textMessage += u'현재 Input Mode: '
                        if  organization[userRequest['user_key']][InputModeString] == 0 :
                            _textMessage += StateButtonList[nx_Child_Sibling(currentState,1,2)][0]
                        elif  organization[userRequest['user_key']][InputModeString] == 1 :
                            _textMessage += StateButtonList[nx_Child_Sibling(currentState,1,2)][1]
                        elif  organization[userRequest['user_key']][InputModeString] == 2 :
                            _textMessage += StateButtonList[nx_Child_Sibling(currentState,1,2)][2]
                        else   :
                            _textMessage += StateButtonList[nx_Child_Sibling(currentState,1,2)][0]
                    return Arrow()._make_Message_Button_change_State( True, _textMessage, True, currentState, nx_Child_Sibling(currentState,1,2) , userRequest ) 
            elif  userRequest['content']  ==  StateButtonList[currentState][4] :  #  go to administrator menu
                return Arrow().make_Message_Button_change_State(currentState, nx_Child_Sibling(currentState,1,4), userRequest)                        
            elif  userRequest['content']  ==  StateButtonList[currentState][5] :  # return to prev menu
                return Arrow().make_Message_Button_change_State( currentState , prev_Parent(currentState,1) , userRequest )                        
            else : 
                _text_ = userRequest['content']+SelectString + u'\n'+ UnderConstructionString
                return Arrow()._make_Message_Button_change_State(True,  _text_, True, currentState,  initial_State , userRequest )                        
        elif   instance[userRequest['user_key']][StateString]  in  \
        [ first_4work_State , first_3work_State , first_3handpiece_State,  first_3com_State, first_4eng_State ] :
            currentState = instance[userRequest['user_key']][StateString]
            if  userRequest['content'].isdigit()  :
                currentContent  = userRequest['content']
                currentIntValue = int( userRequest['content'] ) 
                if  currentIntValue == 0 :
                    return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest)   
                elif instance[userRequest['user_key']][StateString] in [first_4work_State , first_3work_State, first_3handpiece_State] and \
                     currentIntValue  in range(1,96+1) :
                    instance[userRequest['user_key']][SeatNumberString] = currentContent
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,1) , userRequest, request.url_root)
                elif instance[userRequest['user_key']][StateString] in [first_3com_State] and \
                     currentIntValue  in range(1,88+1) :
                    instance[userRequest['user_key']][SeatNumberString] = currentContent
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,1) , userRequest, request.url_root)
                elif instance[userRequest['user_key']][StateString] in [first_4eng_State] and \
                     currentIntValue  in range(1, len_4eng_tables+1) :
                    instance[userRequest['user_key']][SeatNumberString] = currentContent
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child_Sibling(currentState,1,currentIntValue-1) , userRequest, request.url_root)
                else :
                    _textMessage = currentContent+SelectString+u'\n'+ InsertValidNumberString
                    return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState,   currentState , userRequest)
            else : 
                _textMessage = userRequest['content']+SelectString+u'\n'+ InsertNumberString
                return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState, currentState , userRequest)
        # insert device
        elif   instance[userRequest['user_key']][StateString] in \
        [ nx_Child(first_4work_State,1) ,  nx_Child(first_3work_State,1) , nx_Child(first_3handpiece_State,1) , nx_Child(first_3com_State ,1) ] + \
        list( range( nx_Child(first_4eng_State,1) , nx_Child_Sibling(first_4eng_State,1,12-1)+1) ) :
            currentState = instance[userRequest['user_key']][StateString]
            #if userRequest['content'] in  StateButtonList[nx_Child(first_4work_State,1)] :
            if userRequest['content'] in  StateButtonList[currentState] :
                #i = StateButtonList[nx_Child(first_4work_State,1)].index(userRequest['content'])
                i = StateButtonList[currentState].index(userRequest['content'])
                if  i == len(StateButtonList[currentState]) -2 :
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState ,1), userRequest, request.url_root)
                elif i == len(StateButtonList[currentState]) -1 :
                    return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest, request.url_root  )   
                else :
                    instance[userRequest['user_key']][PartString] = userRequest['content']
                    if currentState in range( nx_Child(first_4eng_State,1) , nx_Child_Sibling(first_4eng_State,1,12-1)+1) :
                        instance[userRequest['user_key']][PartString] += u'('+ _4EngDevSNList[currentState][i] +u')'    
                    return Arrow().make_Message_Button_change_State(currentState, nx_Child_Sibling(currentState,2,i) , userRequest )
            else : 
                _textMessage = userRequest['content']+SelectString+u'\n'+'(state:'+ str(instance[userRequest['user_key']][StateString]) + ')'
                instance[userRequest['user_key']] = { StateString : initial_State }            
                return  Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,   initial_State, userRequest)
        # direct insert device
        elif   instance[userRequest['user_key']][StateString] in \
         [ nx_Child(first_4work_State,2) , nx_Child(first_3work_State,2), nx_Child(first_3handpiece_State ,2) , nx_Child(first_3com_State ,2)] + \
         list( range(   nx_Child(nx_Child(first_4eng_State,1),1), nx_Child( nx_Child_Sibling(first_4eng_State,1,12-1), 1)+1 ,  0x10) ) :
            currentState = instance[userRequest['user_key']][StateString]
            if  userRequest['content']  == '0' :
                return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,1) , userRequest, request.url_root)   
            instance[userRequest['user_key']][PartString] = userRequest['content']
            #return Arrow().make_Message_Button_change_State(currentState, nx_Child(currentState,2), userRequest)
            instance[userRequest['user_key']][SymptomString] = u''            
            # no ID info case
            if userRequest['user_key'] not in organization:
                _textMessage = SummaryText()._generate(LastYesNoString+u'\n' ,  temp_organization , instance , userRequest['user_key'])   
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  determineSubGraph(currentState ,5) , userRequest)             
            # already have ID info   
            else : 
                _textMessage = SummaryText()._generate(LastYesNoString + u'\n' ,  organization , instance , userRequest['user_key'])                
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  determineSubGraph(currentState ,5) , userRequest)  
        # insert symptom
        elif  instance[userRequest['user_key']][StateString] in \
        list( range(nx_Child(first_4work_State,3)+0, nx_Child(first_4work_State,3)+ len_4work_part-2 ) )  +  \
        list( range(nx_Child(first_3handpiece_State,3)+0, nx_Child(first_3handpiece_State,3)+ len_3handpiece_part-2 ) ) + \
        list(range(nx_Child(first_3com_State ,3)+0, nx_Child(first_3com_State ,3)+ len_3com_part-2 )) + \
        list(range(nx_Child(first_3work_State,3)+0, nx_Child(first_3work_State ,3)+ len_3work_part-2 )) + \
        _4EngSymptomStateList :

            currentState = instance[userRequest['user_key']][StateString]
            if  userRequest['user_key'] in organization and  InputModeString in  organization[ userRequest['user_key'] ] and \
            organization[ userRequest['user_key'] ][InputModeString] == 1 :  #direct insertion
                if  userRequest['content']  == '0' :
                    return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,2) , userRequest, request.url_root)   
                instance[userRequest['user_key']][SymptomString] = userRequest['content']            
                # no ID info case
                if userRequest['user_key'] not in organization:
                    _textMessage = SummaryText()._generate(LastYesNoString+u'\n' ,  temp_organization , instance , userRequest['user_key'])   
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  determineSubGraph(currentState ,5) , userRequest)             
                # already have ID info   
                else : 
                    _textMessage = SummaryText()._generate(LastYesNoString + u'\n' ,  organization , instance , userRequest['user_key'])                
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  determineSubGraph(currentState ,5) , userRequest)             
            elif  userRequest['user_key'] in organization and  InputModeString in  organization[ userRequest['user_key'] ] and \
            organization[ userRequest['user_key'] ][InputModeString] == 2 :  # mixed insertion

                _textMultiChoice = u'' 
                tokens = re.split(r'(\s*\,\s*|\s+)', userRequest['content'] )
                if  len(tokens) == 1 and \
                    tokens[0].strip().isdigit() and \
                    int ( tokens[0].strip() ) == 0 :
                        return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,2) , userRequest, request.url_root)              
                for i  in  range( 1, len(tokens)-1 )  :
                    if bool (   re.match( r'\s*\,\s*|\s+' , tokens[i] ) ) :
                        if tokens[i-1].strip().isdigit()  :
                            if  bool (   re.match( r'\s*\,\s*|\s+' , tokens[i+1] ) ) :
                                continue 
                            else :
                                tokens[i] =  u','
                        elif tokens[i+1].strip().isdigit() :
                            if  bool (   re.match( r'\s*\,\s*|\s+' , tokens[i-1] ) ) :
                                continue 
                            else :
                                tokens[i] =  u','

                lookup_table = []
                for token in tokens :
                    if token.strip().isdigit() : 
                        _intValue = int(token.strip()) 
                        if _intValue in range( 1, 1+len( StateMultiChoiceList[currentState] ) ) :
                            if  _intValue not in lookup_table :
                                _textMultiChoice += StateMultiChoiceList[currentState][ _intValue-1 ]
                                lookup_table.append( _intValue )
                            else :
                                continue                     
                        else :
                            _textMessage = token.strip()+SelectString+u'\n'+ InsertValidNumberString
                            return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState,   currentState , userRequest)                  
                    else :
                        _textMultiChoice += token

                _MaxSymptomLen = 100 
                TooLongString = u'너무 깁니다'
                if len(_textMultiChoice) > _MaxSymptomLen :
                    _textMessage = _textMultiChoice+ fromStateMessageList[currentState] + TooLongString +u'(' + str(len(_textMultiChoice)) +u'>' +str(_MaxSymptomLen)+u')\n'+ReInsertString
                    return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState,   currentState , userRequest)                  
                instance[userRequest['user_key']][SymptomString] = _textMultiChoice

                _UserRequestKey = userRequest['user_key']
                # ID info in temp_organization 
                if userRequest['user_key'] not in organization:
                    _textMessage = SummaryText()._generate(LastYesNoString+u'\n' ,  temp_organization , instance , userRequest['user_key'])     
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,   determineSubGraph(currentState ,5) , userRequest)             
                #already ID info  in organization  
                else : 
                    _textMessage = SummaryText()._generate(LastYesNoString+u'\n' ,  organization , instance , userRequest['user_key'])                
                    return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,   determineSubGraph(currentState ,5) , userRequest)             

            else :   # button insertion
                if  userRequest['content'] in StateButtonList[ currentState ] :
                    lastKeyIndex = len(StateButtonList[ currentState ])-1
                    #direct describe case
                    if  userRequest['content']  ==  StateButtonList[ currentState ][lastKeyIndex-1] :
                        return Arrow().make_Message_Button_change_State(currentState,  determineSubGraph(currentState ,4), userRequest )
                        #return Arrow().make_Message_Button_change_State(currentState,nx_Child(  determineSubGraph(currentState) ,4), userRequest )
                    elif  userRequest['content']  ==  StateButtonList[ currentState ][lastKeyIndex] :
                        return Arrow().make_Message_Button_change_State(currentState, prev_Parent(currentState,2) , userRequest, request.url_root)   
                    else:
                        instance[userRequest['user_key']][SymptomString] = userRequest['content']            
                        # ID info in temp_organization 
                        if userRequest['user_key'] not in organization:
                            _textMessage = SummaryText()._generate(LastYesNoString+u'\n' ,  temp_organization , instance , userRequest['user_key'])                
                            return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState, determineSubGraph(currentState ,5) , userRequest)             
                            #return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState, nx_Child(  determineSubGraph(currentState) ,5) , userRequest)             
                        #already ID info  in organization  
                        else : 
                            _textMessage = SummaryText()._generate(LastYesNoString+u'\n' ,  organization , instance , userRequest['user_key'])                
                            return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState, determineSubGraph(currentState ,5) , userRequest)                     
                            #return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState, nx_Child(  determineSubGraph(currentState) ,5) , userRequest)                     
                else :
                    _textMessage = userRequest['content']+SelectString+u'\n'
                    instance[userRequest['user_key']] = { StateString : initial_State }            
                    temp = Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  initial_State, userRequest)
                    # have to enable and verify  below code
                    ##instance.pop( userRequest['user_key'] ,  None)
                    return temp
     
        elif   instance[userRequest['user_key']][StateString]  \
        in  [ nx_Child(first_4work_State,4) , nx_Child(first_3work_State,4) , nx_Child(first_3handpiece_State ,4) , nx_Child(first_3com_State ,4) ]  +  \
        list( range(   nx_Child(nx_Child(first_4eng_State,1),3), nx_Child( nx_Child_Sibling(first_4eng_State,1,12-1), 3)+1 ,  0x10) )   :
            currentState = instance[userRequest['user_key']][StateString] 

            if  userRequest['content']  == '0' :
                return Arrow().make_Message_Button_change_State(currentState, restore_prev_State( userRequest['user_key'] )  , userRequest, request.url_root)   
            instance[userRequest['user_key']][SymptomString] = userRequest['content']            
            # no ID info case
            if userRequest['user_key'] not in organization:
                _textMessage = SummaryText()._generate(LastYesNoString+u'\n' ,  temp_organization , instance , userRequest['user_key'])   
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  determineSubGraph(currentState ,5) , userRequest)             
            # already have ID info   
            else : 
                _textMessage = SummaryText()._generate(LastYesNoString + u'\n' ,  organization , instance , userRequest['user_key'])                
                return Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,  determineSubGraph(currentState ,5) , userRequest)             

        # Submit Instance Yes/Yes+/No/Prev
        elif   instance[userRequest['user_key']][StateString]  \
        in [ nx_Child(first_4work_State,5)   ,nx_Child(first_3work_State ,5) , nx_Child(first_3handpiece_State ,5), nx_Child( first_3com_State,5) ] +\
            list( range(   nx_Child(nx_Child(first_4eng_State,1),4), nx_Child( nx_Child_Sibling(first_4eng_State,1,12-1), 4)+1 ,  0x10) ):
            currentState = instance[userRequest['user_key']][StateString]   

            if  userRequest['content']  in [ StateButtonList[ currentState ][0] , StateButtonList[ currentState ][1] ] :

                if  userRequest['user_key'] in temp_organization and \
                    userRequest['user_key'] not in  organization :
                    organization[userRequest['user_key']] = { IDString :  temp_organization[userRequest['user_key']][IDString]    }
                    organization[userRequest['user_key']][NameString] = temp_organization[userRequest['user_key']][NameString]   

                    if  GradeString in temp_organization[userRequest['user_key']]        and  GradeString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][GradeString ] = temp_organization[userRequest['user_key']][GradeString ]   
                    if  RecordedYearString in temp_organization[userRequest['user_key']] and RecordedYearString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][RecordedYearString ] = temp_organization[userRequest['user_key']][RecordedYearString ]   
                    if  InputModeString in temp_organization[userRequest['user_key']]    and InputModeString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][InputModeString] = temp_organization[userRequest['user_key']][InputModeString]   
                    temp_organization.pop( userRequest['user_key'] ,  None)

                _UserRequestKey = userRequest['user_key']
                _Time =  unicode( (datetime.now() + timedelta(hours=time_difference) ).strftime("%Y-%m-%d %H:%M:%S") )
                if _UserRequestKey not in sum_instance    :
                    sum_instance[_UserRequestKey] = []

                _copy_instance = { TimeString:_Time }
                _copy_instance[LocationString] = instance[ _UserRequestKey ][LocationString]
                _copy_instance[SeatNumberString] = instance[ _UserRequestKey ][SeatNumberString]
                _copy_instance[PartString]  = instance[ _UserRequestKey ][PartString]            
                _copy_instance[SymptomString] = instance[ _UserRequestKey ][SymptomString]
                sum_instance[_UserRequestKey].append(_copy_instance)

                if  userRequest['content']  ==  StateButtonList[ currentState ][0] :
                    instance[userRequest['user_key']] = { StateString : initial_State }            
                    return  Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
                else  :
                    return Arrow().make_Message_Button_change_State(currentState,  determineSubGraph(currentState,1)  , userRequest, request.url_root)
            elif userRequest['content']  ==  StateButtonList[ currentState ][3] :  # prev menu
                return Arrow().make_Message_Button_change_State(currentState, restore_prev_State(userRequest['user_key'])  , userRequest)
                #if  userRequest['user_key'] not in temp_organization and \
                #    userRequest['user_key'] in  organization :
                #    return Arrow().make_Message_Button_change_State(currentState, restore_prev_State(userRequest['user_key'])  , userRequest)
                #return Arrow().make_Message_Button_change_State(currentState ,  prev_Parent( currentState, 1 ) , userRequest)             
    #            _textMessage = userRequest['content']+SelectString+u'\n'+ InsertNameString
    #            return Arrow()._make_Message_Button_change_State(_textMessage, False, prev_Parent( currentState, 1 ) , userRequest)             

    #        elif userRequest['content']  ==  StateButtonList[ currentState ][3] :
    #            return Arrow().make_Message_Button_change_State(currentState, nx_Child( determineSubGraph(currentState, True),1)  , userRequest, request.url_root)


            elif  userRequest['content']  in [ StateButtonList[ currentState ][4] , StateButtonList[ currentState ][5] ] :


                if  userRequest['user_key'] in temp_organization and \
                    userRequest['user_key'] not in  organization :
                    organization[userRequest['user_key']] = { IDString :  temp_organization[userRequest['user_key']][IDString]    }
                    organization[userRequest['user_key']][NameString] = temp_organization[userRequest['user_key']][NameString]   

                    if  GradeString in temp_organization[userRequest['user_key']]        and  GradeString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][GradeString ] = temp_organization[userRequest['user_key']][GradeString ]   
                    if  RecordedYearString in temp_organization[userRequest['user_key']] and RecordedYearString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][RecordedYearString ] = temp_organization[userRequest['user_key']][RecordedYearString ]   
                    if  InputModeString in temp_organization[userRequest['user_key']]    and InputModeString not in organization[userRequest['user_key']] :
                        organization[userRequest['user_key']][InputModeString] = temp_organization[userRequest['user_key']][InputModeString]   
                    temp_organization.pop( userRequest['user_key'] ,  None)

                _UserRequestKey = userRequest['user_key']
                _Time =  unicode( (datetime.now() + timedelta(hours=time_difference) ).strftime("%Y-%m-%d %H:%M:%S") )
                if _UserRequestKey not in practice_sum_instance    :
                    practice_sum_instance[_UserRequestKey] = []

                _copy_instance = { TimeString:_Time }
                _copy_instance[LocationString] = instance[ _UserRequestKey ][LocationString]
                _copy_instance[SeatNumberString] = instance[ _UserRequestKey ][SeatNumberString]
                _copy_instance[PartString]  = instance[ _UserRequestKey ][PartString]            
                _copy_instance[SymptomString] = instance[ _UserRequestKey ][SymptomString]
                practice_sum_instance[_UserRequestKey].append(_copy_instance)
                
                if  userRequest['content']  ==  StateButtonList[ currentState ][4] :
                    instance[userRequest['user_key']] = { StateString : initial_State }            
                    return  Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
                else  :
                    return Arrow().make_Message_Button_change_State(currentState,  determineSubGraph(currentState,1)  , userRequest, request.url_root)


            else :   # No case   
                _textMessage = userRequest['content']+ u'\n'+ CancelString 
                instance[userRequest['user_key']] = { StateString : initial_State }            
                return  Arrow()._make_Message_Button_change_State(True, _textMessage, True, currentState,initial_State, userRequest)
           
        #insert Yes of No for delete and prepare branches
        #elif  instance[userRequest['user_key']][StateString] == first_Independent_IDInsert_State+1  :
        elif  instance[userRequest['user_key']][StateString] == nx_Child_Sibling(nx_Child_Sibling(initial_State,1,3),1,1)  :
            currentState = instance[userRequest['user_key']][StateString]              
            if  userRequest['content']  ==  StateButtonList[currentState][0] :
                organization.pop( userRequest['user_key'] ,  None)
                return Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
            elif  userRequest['content']  ==  StateButtonList[currentState][1] :
                return Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
            elif  userRequest['content']  ==  StateButtonList[currentState][2] :
                return Arrow().make_Message_Button_change_State(currentState,  prev_Parent(currentState,1) , userRequest)

        elif  instance[userRequest['user_key']][StateString] == nx_Child_Sibling(nx_Child_Sibling(initial_State,1,3),1,2)  :
            currentState = instance[userRequest['user_key']][StateString]
            if  userRequest['content']  ==  StateButtonList[currentState][0] :
                organization[userRequest['user_key']][InputModeString ] = 0
                return Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
            elif  userRequest['content']  ==  StateButtonList[currentState][1] :
                organization[userRequest['user_key']][InputModeString ] = 1
                return Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
            elif  userRequest['content']  ==  StateButtonList[currentState][2] :
                organization[userRequest['user_key']][InputModeString ] = 2
                return Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
            else :
                return Arrow().make_Message_Button_change_State(currentState, prev_Parent( currentState, 1 ) , userRequest)   

        elif  instance[userRequest['user_key']][StateString] == nx_Child_Sibling(nx_Child_Sibling(initial_State,1,3),1,4)  :
            currentState = instance[userRequest['user_key']][StateString]
            if  os.path.exists(pass_rofile_path)  :
                f = open( pass_rofile_path , 'r')
                p = f.readline().strip()
                f.close()  
                if  userRequest['content'] == p  :
                    _textMessage = userRequest['content']+ fromStateMessageList[currentState]+u'\n'+ toStateMessageList[nx_Child(currentState,1)]+u'\n\n'                
                    ##_textMessage += u'---from memory-----------\n'
                    for key in organization.keys() :
                        _line = key  + u'  '+ str(organization[key][IDString]) + u'  '+ organization[key][NameString]  
                        if GradeString in organization[key].keys() :
                            _line += u'  '+ str(organization[key][GradeString ]) 
                        if RecordedYearString in organization[key].keys() :
                            _line += u'  '+ str(organization[key][RecordedYearString ]) 
                        if InputModeString in organization[key].keys() :
                            _line += u'  '+ str(organization[key][InputModeString ])
                        _line += u'\n'
                        ##_textMessage += _line
                    Org2File(organization, org_rwfile_path)  # organization to organization file
                    _textMessage += u'----------------------\n'
                    _textMessage += u'copy organization\n'
                    _textMessage += u'from memory to storage\n'
                    _textMessage += u'----------------------\n'
                    #_textMessage += SummaryText().showOrgFile()

                    m = MailBodyandAttachment()
                    m.prepare5(True)
                    _textMessage += m.getBody() + u'\n'
                    _textMessage += u'[[practice_sum_instance]]' + u'\n'
                    for key0 in practice_sum_instance :
                        for i in range( len(practice_sum_instance[key0]) ):
                            _textMessage += SummaryText()._generate(u'---------' + str(i+1) +  u'------------\n' , organization, practice_sum_instance , key0 , i)

                    subject = u'전체고장 확인(총 '+str(m.getInstanceCount())+u'건)('    
                    subject += unicode ( (datetime.now() + timedelta(hours=time_difference)  ).strftime("%Y-%m-%d")  )+u')'

                    parameter_list = [ emailAdminList , subject , ( unicode(request.url_root)+u'\n'+SummaryText().showOrgFile()+u'\n'+m.getBody()).encode('utf-8') , m.getAttachmentList()  ] 
                    Timer(1.0, mail, parameter_list).start()

                    return Arrow()._make_Message_Button_change_State(True, _textMessage,  False, currentState, nx_Child(currentState,1) , userRequest)     
                else :
                    return Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
            else :
                return Arrow().make_Message_Button_change_State(currentState, initial_State, userRequest)
        elif  instance[userRequest['user_key']][StateString] == nx_Child(nx_Child_Sibling(nx_Child_Sibling(initial_State,1,3),1,4),1) :
            currentState = instance[userRequest['user_key']][StateString]
            if  len(userRequest['content'])  > 0  :
                return Arrow().make_Message_Button_change_State( currentState, initial_State, userRequest  )
            else :
                return Arrow().make_Message_Button_change_State( currentState, initial_State, userRequest )

        #else:     
        _text = '(state:'+ format(instance[userRequest['user_key']][StateString] , '#04x' ) + ')(content:' + userRequest['content'] + ')'
        text = "Invalid State!"   
        text = text + "!! (" + _text + ")"  
        text = text + '(user_key:' + userRequest['user_key'] + ')'

        if userRequest['user_key'] not in instance :
            text = text + 'user_key is NOT in instance'
        else :
            text = text + 'user_key is in instance'
        textContent = {"text":text}
        textMessage = {"message":textContent}
        return jsonify(textMessage)
    except Exception as ex :
        return  Arrow()._make_Message_Button_change_State(True,  traceback.format_exc()  , True,  currentState,initial_State, userRequest)
 #               return  Arrow()._make_Message_Button_change_State(True,  str(ex) , True,  currentState,initial_State, userRequest)


@app.errorhandler(404)
def page_not_found(e):
    text = "Invalid command!"
    textContent = {"text":text}
    errorMessage = {"message":textContent}
    return jsonify(errorMessage)

port = os.getenv('PORT', '5000')
if __name__ == "__main__":
	app.run(host='0.0.0.0', port=int(port))
